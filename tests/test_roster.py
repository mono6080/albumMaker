# 名冊與學期彙整匯出測試
# 覆蓋：園所目前名單身分、相本快照不可重連，以及學期匯出預覽與 ZIP。

from io import BytesIO
from types import SimpleNamespace
from zipfile import ZipFile

from fastapi.testclient import TestClient
from sqlalchemy import text

from tests.helpers import (
    assert_status,
    create_project,
    create_template_with_page,
    create_user,
    jpeg_bytes,
    login,
    revisioned_project_url,
    smoke_layout,
    started_client,
    unique_name,
    use_tmp_uploads,
)

from database import (
    Semester,
    Project,
    SessionLocal,
    ProjectStudent,
    Template,
    engine,
)
from services import semester_export_service, semester_render_service
from services.output_keys import get_student_pdf_key
from services.roster_identity_service import normalize_child_name
from services.storage_factory import get_storage
from services.student_progress import summarize_student_progress


def create_active_period(client: TestClient, department: str = "infant") -> dict:
    response = client.post(
        "/api/templates/periods",
        data={"name": unique_name("period"), "department": department, "status": "active"},
    )
    assert_status(response, 200)
    return response.json()


def create_scoped_classroom(
    client: TestClient,
    *,
    campus_id: int | None = None,
    department: str = "infant",
) -> tuple[int, int]:
    if campus_id is None:
        campus_response = client.post(
            "/api/organization/campuses",
            json={"name": unique_name("scope_campus")},
        )
        assert_status(campus_response, 201)
        campus_id = campus_response.json()["id"]
    classroom_response = client.post(
        "/api/organization/classrooms",
        json={
            "campus_id": campus_id,
            "department": department,
            "name": unique_name(f"scope_{department}"),
        },
    )
    assert_status(classroom_response, 201)
    return campus_id, classroom_response.json()["id"]


def set_campus_supervisor_scope(
    client: TestClient,
    campus_id: int,
    supervisor_id: int,
    *,
    department: str | None = None,
) -> None:
    campus_supervisor_ids = [supervisor_id] if department is None else []
    department_supervisors = [
        {
            "department": department_name,
            "supervisor_ids": (
                [supervisor_id] if department == department_name else []
            ),
        }
        for department_name in ("infant", "academy")
    ]
    response = client.put(
        f"/api/organization/campuses/{campus_id}/supervisors",
        json={
            "campus_supervisor_ids": campus_supervisor_ids,
            "department_supervisors": department_supervisors,
        },
    )
    assert_status(response, 200)


def set_classroom_teachers(
    client: TestClient,
    classroom_id: int,
    teacher_ids: list[int],
) -> None:
    response = client.put(
        f"/api/organization/classrooms/{classroom_id}/teachers",
        json={
            "teachers": [
                {
                    "teacher_id": teacher_id,
                    "duty": "lead" if index == 0 else "co_teacher",
                }
                for index, teacher_id in enumerate(teacher_ids)
            ]
        },
    )
    assert_status(response, 200)


def add_classroom_members(
    client: TestClient,
    classroom_id: int,
    names: list[str],
) -> None:
    response = client.post(
        f"/api/organization/classrooms/{classroom_id}/members/batch",
        json={"members": [{"name": name} for name in names]},
    )
    assert_status(response, 201)


def create_classroom_project(
    client: TestClient,
    classroom_id: int,
    template_id: int,
) -> int:
    db = SessionLocal()
    try:
        template_period_id = db.get(Template, template_id).period_id
    finally:
        db.close()
    overview = client.get("/api/organization/overview")
    assert_status(overview, 200)
    work_slot_id = next(
        slot["id"]
        for slot in overview.json()["work_slots"]
        if slot["classroom_id"] == classroom_id
        and slot["template_period_id"] == template_period_id
    )
    response = client.post(
        f"/api/organization/classrooms/{classroom_id}/projects",
        json={
            "name": unique_name("class_project"),
            "template_id": template_id,
            "work_slot_id": work_slot_id,
        },
    )
    assert_status(response, 201)
    return response.json()["id"]


def create_unassigned_migration_project(
    template_id: int,
    owner_id: int,
    student_name: str,
) -> int:
    """直接種入待 migration 舊資料；正常 API 已禁止建立未歸班專案。"""
    db = SessionLocal()
    try:
        template = db.get(Template, template_id)
        project = Project(
            name=unique_name("migration_required"),
            template_id=template.id,
            template_revision=template.revision,
            department=template.period.department,
            template_period_id=template.period.id,
            owner_id=owner_id,
        )
        db.add(project)
        db.flush()
        db.add(ProjectStudent(project_id=project.id, name=student_name, order_index=0))
        db.commit()
        return project.id
    finally:
        db.close()


def create_period_template_project(
    client: TestClient,
    period_id: int,
    student_names: list[str] | None = None,
) -> int:
    """建立掛在指定期別下的模板與班級專案，回傳 project_id。"""
    template_id, _ = create_template_with_page(
        client,
        period_id=period_id,
    )
    return create_project(
        client,
        template_id,
        student_names=student_names,
    )


def add_students(client: TestClient, project_id: int, names: list[str]) -> dict[str, int]:
    """讀取建立專案時由班級目前名單產生的學生快照。"""
    detail = client.get(f"/api/projects/{project_id}")
    assert_status(detail, 200)
    students_by_identity = {
        normalize_child_name(student["name"]): student["id"]
        for student in detail.json()["students"]
    }
    assert {
        normalize_child_name(name) for name in names
    } <= set(students_by_identity)
    return {
        name: students_by_identity[normalize_child_name(name)]
        for name in names
    }


def roster_child_id_of(student_id: int) -> int | None:
    db = SessionLocal()
    try:
        return db.query(ProjectStudent).filter(ProjectStudent.id == student_id).one().roster_child_id
    finally:
        db.close()


def assigned_identity_anomaly_count(client: TestClient) -> int:
    overview = client.get("/api/organization/overview")
    assert_status(overview, 200)
    return overview.json()["migration_status"]["assigned_identity_anomaly_count"]


def reporting_semester_id(client: TestClient, period_ids: list[int]) -> int:
    response = client.get("/api/roster/semesters")
    assert_status(response, 200)
    requested_ids = set(period_ids)
    return next(
        term["id"]
        for term in response.json()["terms"]
        if requested_ids <= {
            period["template_period_id"] for period in term["periods"]
        }
    )


def get_semester_preview(client: TestClient, period_ids: list[int]):
    return client.get(
        "/api/roster/semester-export",
        params={
            "semester_id": reporting_semester_id(client, period_ids),
            "period_ids": period_ids,
        },
    )


def preview_children(preview_payload: dict) -> list[dict]:
    return [
        child
        for classroom_group in preview_payload["classroom_groups"]
        for child in classroom_group["children"]
    ]


def child_entries(child: dict) -> list[dict]:
    return [entry for cell in child["cells"] for entry in cell["entries"]]


def test_current_roster_identity_groups_multiple_period_snapshots():
    with started_client() as client:
        login(client)
        teacher, _ = create_user(client, "teacher")
        period_a = create_active_period(client)
        period_b = create_active_period(client)
        template_a_id, _ = create_template_with_page(client, period_id=period_a["id"])
        template_b_id, _ = create_template_with_page(client, period_id=period_b["id"])
        _, classroom_id = create_scoped_classroom(client)
        set_classroom_teachers(client, classroom_id, [teacher["id"]])
        add_classroom_members(client, classroom_id, ["王小明", "李小華"])
        project_a = create_classroom_project(client, classroom_id, template_a_id)
        project_b = create_classroom_project(client, classroom_id, template_b_id)

        students_a = add_students(client, project_a, ["王小明", "李小華"])
        students_b = add_students(client, project_b, ["王小明"])

        ming_a = roster_child_id_of(students_a["王小明"])
        ming_b = roster_child_id_of(students_b["王小明"])
        hua_a = roster_child_id_of(students_a["李小華"])
        assert ming_a is not None
        assert ming_a == ming_b
        assert hua_a is not None and hua_a != ming_a

        # 匯出預覽依名冊孩子分組
        preview = get_semester_preview(client, [period_a["id"], period_b["id"]])
        assert_status(preview, 200)
        preview_data = preview.json()
        assert [period["id"] for period in preview_data["periods"]] == [period_a["id"], period_b["id"]]
        groups_by_name = {group["name"]: group for group in preview_children(preview_data)}
        assert len(child_entries(groups_by_name["王小明"])) == 2
        assert len(child_entries(groups_by_name["李小華"])) == 2
        assert preview_data["unlinked"] == []
        assert all(entry["has_pdf"] is False for entry in child_entries(groups_by_name["王小明"]))


def test_new_entrants_get_new_identity_and_same_class_name_is_skipped():
    with started_client() as client:
        login(client)
        campus_id, classroom_a_id = create_scoped_classroom(client)
        _, classroom_b_id = create_scoped_classroom(client, campus_id=campus_id)

        first_response = client.post(
            f"/api/organization/classrooms/{classroom_a_id}/members/batch",
            json={"members": [{"name": "王 小明"}, {"name": "王小明"}]},
        )
        assert_status(first_response, 201)
        assert len(first_response.json()["created"]) == 1
        assert first_response.json()["skipped"] == ["王小明"]

        second_response = client.post(
            f"/api/organization/classrooms/{classroom_b_id}/members/batch",
            json={"members": [{"name": "王小明"}]},
        )
        assert_status(second_response, 201)
        first_child_id = first_response.json()["created"][0]["roster_child_id"]
        second_child_id = second_response.json()["created"][0]["roster_child_id"]
        assert first_child_id != second_child_id


def test_roster_mutation_routes_are_absent_and_snapshots_stay_unchanged():
    with started_client() as client:
        login(client)
        period = create_active_period(client)
        project = create_period_template_project(client, period["id"], ["王小明"])
        student_id = add_students(client, project, ["王小明"])["王小明"]
        original_child_id = roster_child_id_of(student_id)

        openapi_paths = client.get("/openapi.json").json()["paths"]
        assert "/api/roster/students/{student_id}/link" not in openapi_paths
        assert "/api/roster/children/{child_id}/merge/{target_child_id}" not in openapi_paths

        link_response = client.put(
            f"/api/roster/students/{student_id}/link",
            json={"create_new": True},
        )
        assert_status(link_response, 405)
        merge_response = client.post(
            f"/api/roster/children/{original_child_id}/merge/{original_child_id}"
        )
        assert_status(merge_response, 405)
        assert roster_child_id_of(student_id) == original_child_id


def test_semester_preview_reports_legacy_identity_anomaly_without_candidates():
    with started_client() as client:
        login(client)
        period = create_active_period(client)
        project = create_period_template_project(client, period["id"], ["待移除快照"])

        db = SessionLocal()
        try:
            db.query(ProjectStudent).filter(ProjectStudent.project_id == project).delete()
            # 直接種入 migration 前既存的異常快照；正式流程不可再把已歸班身分改回 NULL。
            student = ProjectStudent(
                project_id=project,
                name="舊資料學生",
                order_index=0,
            )
            db.add(student)
            db.commit()
            student_id = student.id
        finally:
            db.close()

        preview = get_semester_preview(client, [period["id"]])
        assert_status(preview, 200)
        assert "舊資料學生" not in {
            child["name"] for child in preview_children(preview.json())
        }
        assert len(preview.json()["unlinked"]) == 1
        anomaly = preview.json()["unlinked"][0]
        assert anomaly["student_id"] == student_id
        assert anomaly["student_name"] == "舊資料學生"
        assert "candidates" not in anomaly


def test_semester_export_lists_invalid_child_fk_without_grouping_or_download(
    monkeypatch,
    tmp_path,
):
    use_tmp_uploads(monkeypatch, tmp_path)
    with started_client() as client:
        login(client)
        period = create_active_period(client)
        project_id = create_period_template_project(client, period["id"], ["待移除快照"])
        anomaly_count_before = assigned_identity_anomaly_count(client)

        db = SessionLocal()
        try:
            db.query(ProjectStudent).filter(ProjectStudent.project_id == project_id).delete()
            db.commit()
        finally:
            db.close()

        raw_connection = engine.raw_connection()
        cursor = raw_connection.cursor()
        cursor.execute("SELECT COALESCE(MAX(id), 0) + 1000000 FROM students")
        invalid_child_id = cursor.fetchone()[0]
        cursor.execute("PRAGMA foreign_keys = OFF")
        cursor.execute(
            """
            INSERT INTO project_students (
                project_id, name, order_index, pages_data_json, roster_child_id
            ) VALUES (?, ?, ?, ?, ?)
            """,
            (project_id, "外鍵遺失學生", 0, "[]", invalid_child_id),
        )
        student_id = cursor.lastrowid
        output_key = get_student_pdf_key(project_id, student_id)
        cursor.execute(
            "UPDATE project_students SET output_filename = ? WHERE id = ?",
            (output_key, student_id),
        )
        raw_connection.commit()
        cursor.execute("PRAGMA foreign_keys = ON")
        raw_connection.commit()
        cursor.close()
        raw_connection.close()
        get_storage().put(output_key, b"%PDF-1.4\n%%EOF")

        try:
            preview = get_semester_preview(client, [period["id"]])
            assert_status(preview, 200)
            preview_data = preview.json()
            assert "外鍵遺失學生" not in {
                child["name"] for child in preview_children(preview_data)
            }
            assert len(preview_data["unlinked"]) == 1
            anomaly = preview_data["unlinked"][0]
            assert anomaly["student_id"] == student_id
            assert anomaly["identity_anomalies"] == ["invalid_roster_child"]
            assert assigned_identity_anomaly_count(client) == anomaly_count_before + 1

            download = client.get(
                "/api/roster/semester-export/download",
                params={
                    "semester_id": reporting_semester_id(client, [period["id"]]),
                    "period_ids": [period["id"]],
                    "mode": "print",
                },
            )
            assert_status(download, 200)
            with ZipFile(BytesIO(download.content)) as zip_archive:
                assert zip_archive.namelist() == ["匯出說明.txt"]
                manifest = zip_archive.read("匯出說明.txt").decode("utf-8")
            assert "外鍵遺失學生" in manifest
            assert "身分異常" in manifest
        finally:
            db = SessionLocal()
            try:
                db.query(ProjectStudent).filter(ProjectStudent.id == student_id).delete()
                db.commit()
            finally:
                db.close()


def test_duplicate_child_identity_is_listed_and_skipped_by_missing_render(
    monkeypatch,
    tmp_path,
):
    use_tmp_uploads(monkeypatch, tmp_path)
    with started_client() as client:
        login(client)
        period = create_active_period(client)
        project_id = create_period_template_project(
            client,
            period["id"],
            ["正常學生", "重複來源"],
        )
        student_ids = add_students(client, project_id, ["正常學生", "重複來源"])
        anomaly_count_before = assigned_identity_anomaly_count(client)

        db = SessionLocal()
        try:
            source_student = db.get(ProjectStudent, student_ids["重複來源"])
            duplicate_student = ProjectStudent(
                project_id=project_id,
                name="重複影本",
                order_index=2,
                pages_data_json="[]",
                roster_child_id=source_student.roster_child_id,
            )
            db.add(duplicate_student)
            db.commit()
            duplicate_student_id = duplicate_student.id
        finally:
            db.close()

        try:
            preview = get_semester_preview(client, [period["id"]])
            assert_status(preview, 200)
            preview_data = preview.json()
            grouped_names = {
                group["name"] for group in preview_children(preview_data)
            }
            assert "正常學生" in grouped_names
            assert "重複來源" in grouped_names
            assert "重複影本" not in grouped_names
            duplicate_child = next(
                child
                for child in preview_children(preview_data)
                if child["name"] == "重複來源"
            )
            assert duplicate_child["cells"][0]["status"] == "no_album"
            assert duplicate_child["cells"][0]["entries"] == []
            anomalies_by_name = {
                entry["student_name"]: entry["identity_anomalies"]
                for entry in preview_data["unlinked"]
            }
            assert anomalies_by_name == {
                "重複來源": ["duplicate_project_roster_child"],
                "重複影本": ["duplicate_project_roster_child"],
            }
            assert assigned_identity_anomaly_count(client) == anomaly_count_before + 2

            rendered_students = []

            def render_one(project, student, project_id, db):
                rendered_students.append(student.name)
                return {"pdf": "unused", "pages": 1}

            monkeypatch.setattr(
                semester_render_service,
                "render_and_save_student_album",
                render_one,
            )
            render_db = SessionLocal()
            try:
                result = semester_render_service.render_missing_semester_albums(
                    render_db,
                    reporting_semester_id(client, [period["id"]]),
                    [period["id"]],
                )
            finally:
                render_db.close()
            assert rendered_students == ["正常學生"]
            assert result == {"rendered": 1, "errors": []}
        finally:
            db = SessionLocal()
            try:
                db.query(ProjectStudent).filter(ProjectStudent.id == duplicate_student_id).delete()
                db.commit()
            finally:
                db.close()


def test_roster_read_endpoints_require_admin():
    with started_client() as client:
        login(client)
        teacher, teacher_password = create_user(client, "teacher")
        semester_id = reporting_semester_id(client, [1])

        client.cookies.clear()
        login(client, teacher["username"], teacher_password)
        preview = client.get(
            "/api/roster/semester-export",
            params={"semester_id": semester_id, "period_ids": [1]},
        )
        assert_status(preview, 403)


def test_supervisor_scoped_preview_and_no_export():
    with started_client() as client:
        login(client)
        supervisor, supervisor_password = create_user(client, "supervisor")
        managed_teacher, _ = create_user(client, "teacher")
        db = SessionLocal()
        try:
            db.execute(
                text("""
                    INSERT INTO legacy_teacher_supervisor_links (
                        teacher_id, supervisor_id,
                        teacher_name_snapshot, supervisor_name_snapshot
                    ) VALUES (
                        :teacher_id, :supervisor_id,
                        :teacher_name, :supervisor_name
                    )
                """),
                {
                    "teacher_id": managed_teacher["id"],
                    "supervisor_id": supervisor["id"],
                    "teacher_name": managed_teacher["display_name"],
                    "supervisor_name": supervisor["display_name"],
                },
            )
            db.commit()
        finally:
            db.close()
        period = create_active_period(client)
        template_id, _ = create_template_with_page(client, period_id=period["id"])

        # 封存關係與 owner 都不授權待 migration 的未歸班專案。
        create_unassigned_migration_project(
            template_id,
            managed_teacher["id"],
            "舊未歸班學生",
        )

        # 園所 scope 內的班級相本才會進主管預覽。
        campus_id, classroom_id = create_scoped_classroom(client)
        set_campus_supervisor_scope(client, campus_id, supervisor["id"])
        set_classroom_teachers(client, classroom_id, [managed_teacher["id"]])
        add_classroom_members(client, classroom_id, ["老師的學生"])
        create_classroom_project(client, classroom_id, template_id)

        # 主管：只看得到園所 scope 內的班級專案。
        client.cookies.clear()
        login(client, supervisor["username"], supervisor_password)
        semester_id = reporting_semester_id(client, [period["id"]])
        preview = get_semester_preview(client, [period["id"]])
        assert_status(preview, 200)
        child_names = {group["name"] for group in preview_children(preview.json())}
        assert "老師的學生" in child_names
        assert "舊未歸班學生" not in child_names
        assert not any(
            row["student_name"] == "舊未歸班學生"
            for row in preview.json()["unlinked"]
        )

        # 匯出與補渲染對主管一律 403
        download = client.get(
            "/api/roster/semester-export/download",
            params={
                "semester_id": semester_id,
                "period_ids": [period["id"]],
            },
        )
        assert_status(download, 403)
        render_missing = client.post(
            "/api/roster/semester-export/render-missing",
            json={
                "semester_id": semester_id,
                "period_ids": [period["id"]],
            },
        )
        assert_status(render_missing, 403)

        db = SessionLocal()
        try:
            db.execute(
                text("""
                    DELETE FROM legacy_teacher_supervisor_links
                    WHERE teacher_id = :teacher_id AND supervisor_id = :supervisor_id
                """),
                {
                    "teacher_id": managed_teacher["id"],
                    "supervisor_id": supervisor["id"],
                },
            )
            db.commit()
        finally:
            db.close()


def test_supervisor_reporting_uses_union_of_organization_scopes_only():
    from openpyxl import load_workbook

    with started_client() as client:
        login(client)
        supervisor, supervisor_password = create_user(client, "supervisor")
        teachers = [create_user(client, "teacher")[0] for _ in range(5)]
        infant_period = create_active_period(client, "infant")
        academy_period = create_active_period(client, "academy")
        infant_template_id, _ = create_template_with_page(
            client,
            period_id=infant_period["id"],
        )
        academy_template_id, _ = create_template_with_page(
            client,
            period_id=academy_period["id"],
        )

        campus_a_id, campus_a_infant_id = create_scoped_classroom(client)
        _, campus_a_academy_id = create_scoped_classroom(
            client,
            campus_id=campus_a_id,
            department="academy",
        )
        campus_b_id, campus_b_infant_id = create_scoped_classroom(client)
        _, campus_b_academy_id = create_scoped_classroom(
            client,
            campus_id=campus_b_id,
            department="academy",
        )
        set_campus_supervisor_scope(client, campus_a_id, supervisor["id"])
        set_campus_supervisor_scope(
            client,
            campus_b_id,
            supervisor["id"],
            department="infant",
        )
        db = SessionLocal()
        try:
            current_term = db.query(Semester).filter(
                Semester.status.in_(("imported", "active"))
            ).one()
            current_term.status = "active"
            db.commit()
        finally:
            db.close()
        for classroom_id, teacher in zip(
            (
                campus_a_infant_id,
                campus_a_academy_id,
                campus_b_infant_id,
                campus_b_academy_id,
            ),
            teachers[:4],
            strict=True,
        ):
            set_classroom_teachers(client, classroom_id, [teacher["id"]])

        add_classroom_members(client, campus_a_infant_id, ["全校範圍學生"])
        visible_project_id = create_classroom_project(
            client,
            campus_a_infant_id,
            infant_template_id,
        )
        # 專案仍歸建立當時 owner；學期中途換老師時，兩位都是這個班本學期的老師。
        set_classroom_teachers(client, campus_a_infant_id, [teachers[4]["id"]])
        add_classroom_members(client, campus_b_academy_id, ["部門外學生"])
        hidden_project_id = create_classroom_project(
            client,
            campus_b_academy_id,
            academy_template_id,
        )
        legacy_project_id = create_unassigned_migration_project(
            infant_template_id,
            teachers[3]["id"],
            "未歸班學生",
        )

        client.cookies.clear()
        login(client, supervisor["username"], supervisor_password)
        classrooms = client.get("/api/organization/my-classrooms")
        assert_status(classrooms, 200)
        assert {row["id"] for row in classrooms.json()["classrooms"]} == {
            campus_a_infant_id,
            campus_a_academy_id,
            campus_b_infant_id,
        }

        period_ids = [infant_period["id"], academy_period["id"]]
        semester_id = reporting_semester_id(client, period_ids)
        progress = client.get(
            "/api/roster/teacher-progress",
            params={"semester_id": semester_id},
        )
        assert_status(progress, 200)
        classroom_rows = {
            row["classroom_id"]: row for row in progress.json()["classrooms"]
        }
        assert set(classroom_rows) == {
            campus_a_infant_id,
            campus_a_academy_id,
            campus_b_infant_id,
        }
        scoped_project_ids = {
            project["project_id"]
            for row in classroom_rows.values()
            for slot in row["slots"]
            for project in slot["projects"]
        }
        assert scoped_project_ids == {visible_project_id}
        assert hidden_project_id not in scoped_project_ids
        assert legacy_project_id not in scoped_project_ids
        assert {
            teacher["user_id"]: teacher["ended_at"] is None
            for teacher in classroom_rows[campus_a_infant_id]["teachers"]
        } == {teachers[0]["id"]: False, teachers[4]["id"]: True}
        assert all(
            not slot["projects"]
            for classroom_id in (campus_a_academy_id, campus_b_infant_id)
            for slot in classroom_rows[classroom_id]["slots"]
        )

        preview = get_semester_preview(client, period_ids)
        assert_status(preview, 200)
        assert {group["name"] for group in preview_children(preview.json())} == {
            "全校範圍學生"
        }

        workbook_response = client.get(
            "/api/roster/teacher-overview/export",
            params={"semester_id": semester_id},
        )
        assert_status(workbook_response, 200)
        workbook = load_workbook(BytesIO(workbook_response.content))
        assert workbook.sheetnames == ["摘要", "班級期別", "學生明細"]
        detail_student_names = {
            row[7]
            for row in workbook["學生明細"].iter_rows(min_row=2, values_only=True)
        }
        assert detail_student_names == {"全校範圍學生"}


def test_teacher_overview_excel_export():
    from openpyxl import load_workbook

    with started_client() as client:
        login(client)
        teacher, _ = create_user(client, "teacher")
        period = create_active_period(client)
        template_id, _ = create_template_with_page(client, period_id=period["id"])
        campus_id, classroom_id = create_scoped_classroom(client)
        set_classroom_teachers(client, classroom_id, [teacher["id"]])
        student_names = [unique_name("excel_student_a"), unique_name("excel_student_b")]
        add_classroom_members(client, classroom_id, student_names)
        create_classroom_project(client, classroom_id, template_id)
        semester_id = reporting_semester_id(client, [period["id"]])

        export = client.get(
            "/api/roster/teacher-overview/export",
            params={
                "semester_id": semester_id,
                "campus_id": campus_id,
                "classroom_id": classroom_id,
            },
        )
        assert_status(export, 200)
        assert export.headers["content-type"].startswith("application/vnd.openxmlformats")

        workbook = load_workbook(BytesIO(export.content))
        assert workbook.sheetnames == ["摘要", "班級期別", "學生明細"]
        summary_rows = list(workbook["摘要"].iter_rows(values_only=True))
        assert summary_rows[0] == ("項目", "數量")
        assert ("專案數", 1) in summary_rows

        slot_rows = list(workbook["班級期別"].iter_rows(values_only=True))
        project_slot_row = next(row for row in slot_rows[1:] if row[7] == 1)
        assert project_slot_row[10] == 2
        detail_rows = list(workbook["學生明細"].iter_rows(values_only=True))
        exported_student_names = {row[7] for row in detail_rows[1:]}
        assert set(student_names) <= exported_student_names
        assert all(row[0] and row[2] for row in detail_rows[1:])


def test_teacher_progress_includes_idle_teachers_and_photo_counts(monkeypatch, tmp_path):
    use_tmp_uploads(monkeypatch, tmp_path)

    with started_client() as client:
        login(client)
        supervisor, supervisor_password = create_user(client, "supervisor")
        managed_teacher, teacher_password = create_user(client, "teacher")
        idle_teacher, _ = create_user(client, "teacher")
        unassigned_teacher, _ = create_user(client, "teacher")
        period = create_active_period(client)
        template_id, _ = create_template_with_page(client, period_id=period["id"])
        campus_id, classroom_id = create_scoped_classroom(client)
        set_campus_supervisor_scope(client, campus_id, supervisor["id"])
        set_classroom_teachers(
            client,
            classroom_id,
            [managed_teacher["id"], idle_teacher["id"]],
        )
        first_student_name = unique_name("progress_student_a")
        second_student_name = unique_name("progress_student_b")
        add_classroom_members(
            client,
            classroom_id,
            [first_student_name, second_student_name],
        )
        teacher_project = create_classroom_project(
            client,
            classroom_id,
            template_id,
        )

        # 管轄老師的專案：兩位學生、只有一格照片被填
        client.cookies.clear()
        login(client, managed_teacher["username"], teacher_password)
        project_detail = client.get(f"/api/projects/{teacher_project}")
        assert_status(project_detail, 200)
        students = {
            student["name"]: student["id"]
            for student in project_detail.json()["students"]
        }
        photo = client.post(
            revisioned_project_url(
                client,
                teacher_project,
                f"/api/projects/{teacher_project}/students/{students[first_student_name]}/pages/0/photos/1",
            ),
            files={"file": ("smoke.jpg", jpeg_bytes(), "image/jpeg")},
        )
        assert_status(photo, 200)

        # 主管視角：只看到園所 scope 的目前老師（含還沒開工的），看不到 admin 專案。
        client.cookies.clear()
        login(client, supervisor["username"], supervisor_password)
        semester_id = reporting_semester_id(client, [period["id"]])
        progress = client.get(
            "/api/roster/teacher-progress",
            params={"semester_id": semester_id},
        )
        assert_status(progress, 200)
        classroom = progress.json()["classrooms"][0]
        assert classroom["classroom_id"] == classroom_id
        assert {teacher["user_id"] for teacher in classroom["teachers"]} == {
            managed_teacher["id"],
            idle_teacher["id"],
        }
        student_names = {
            student["student_name"]
            for slot in classroom["slots"]
            for project in slot["projects"]
            for student in project["students"]
        }
        assert student_names == {first_student_name, second_student_name}

        slot = next(slot for slot in classroom["slots"] if slot["projects"])
        assert slot["creation_status"] == "single"
        project_progress = slot["projects"][0]
        # smoke_layout 每頁 1 照片格與 1 個可填文字；模板範例字不算老師已填。
        assert project_progress["photo_total"] == 2
        assert project_progress["photo_filled"] == 1
        assert project_progress["text_total"] == 2
        assert project_progress["text_filled"] == 0
        assert project_progress["blank_text_count"] == 2
        progress_by_student = {
            student["student_name"]: student for student in project_progress["students"]
        }
        assert progress_by_student[first_student_name]["photo_filled"] == 1
        assert progress_by_student[second_student_name]["photo_filled"] == 0
        assert progress_by_student[first_student_name]["text_total"] == 1
        assert progress_by_student[first_student_name]["text_filled"] == 0

        # admin 也是依全園目前編制展開，不把只有 teacher 角色但未編班的帳號塞進進度。
        client.cookies.clear()
        login(client)
        admin_progress = client.get(
            "/api/roster/teacher-progress",
            params={"semester_id": semester_id},
        )
        assert_status(admin_progress, 200)
        admin_teacher_ids = {
            teacher["user_id"]
            for classroom in admin_progress.json()["classrooms"]
            for teacher in classroom["teachers"]
        }
        assert managed_teacher["id"] in admin_teacher_ids
        assert idle_teacher["id"] in admin_teacher_ids
        assert unassigned_teacher["id"] not in admin_teacher_ids

        # 老師本人無權查看
        client.cookies.clear()
        login(client, managed_teacher["username"], teacher_password)
        forbidden = client.get(
            "/api/roster/teacher-progress",
            params={"semester_id": semester_id},
        )
        assert_status(forbidden, 403)


def test_teacher_progress_ignores_hidden_group_photos_and_texts():
    layout = smoke_layout()
    layout["photo_slots"].append({
        "id": 2,
        "x": 360,
        "y": 96,
        "width": 240,
        "height": 180,
    })
    layout["text_labels"].append({
        "id": 2,
        "x": 96,
        "y": 480,
        "width": 360,
        "height": 96,
        "text": "",
    })
    layout["text_labels"].append({
        "id": 3,
        "x": 96,
        "y": 600,
        "width": 360,
        "height": 96,
        "text": "固定標題",
        "text_role": "static",
    })
    layout["group_contract"] = "nested-world-v2"
    layout["groups"] = [{
        "id": "hidden-progress",
        "z_index": 10,
        "selection_rotation": 0,
        "visible": False,
        "children": [
            {"type": "photo", "id": 2},
            {"type": "text", "id": 2},
        ],
    }]

    result = summarize_student_progress(
        [{
            "page_index": 0,
            "photos": {"1": "visible.jpg", "2": "hidden.jpg"},
            "label_texts": {"2": ""},
        }],
        [layout],
        {"0": {"1": "老師已填"}},
    )

    assert result == (1, 1, 1, 1)


def test_teacher_progress_combines_class_and_individual_text_coverage():
    layout = smoke_layout()
    layout["photo_slots"] = []
    layout["text_labels"] = [
        {
            "id": label_id,
            "x": 40,
            "y": label_id * 50,
            "width": 300,
            "height": 40,
            "text": f"模板範例 {label_id}",
        }
        for label_id in range(1, 13)
    ]
    project_label_texts = {
        "0": {
            str(label_id): f"全班文字 {label_id}"
            for label_id in range(1, 12)
        }
    }

    first_student = summarize_student_progress(
        [{"page_index": 0, "label_texts": {"12": "甲的個人文字"}}],
        [layout],
        project_label_texts,
    )
    second_student = summarize_student_progress(
        [{"page_index": 0, "label_texts": {"12": "乙的個人文字"}}],
        [layout],
        project_label_texts,
    )
    missing_second_student = summarize_student_progress(
        [{"page_index": 0, "label_texts": {}}],
        [layout],
        project_label_texts,
    )

    assert first_student == (0, 0, 12, 12)
    assert second_student == (0, 0, 12, 12)
    assert (
        first_student[2] + second_student[2],
        first_student[3] + second_student[3],
    ) == (24, 24)
    assert (
        first_student[2] + missing_second_student[2],
        first_student[3] + missing_second_student[3],
    ) == (23, 24)


def start_and_wait_render_job(client: TestClient, period_ids: list[int], timeout_seconds: float = 60) -> dict:
    """啟動補渲染 job 並輪詢到結束，回傳最終 job 狀態。"""
    import time

    start = client.post(
        "/api/roster/semester-export/render-missing",
        json={
            "semester_id": reporting_semester_id(client, period_ids),
            "period_ids": period_ids,
        },
    )
    assert_status(start, 200)
    job_id = start.json()["job_id"]
    deadline = time.time() + timeout_seconds
    while time.time() < deadline:
        progress = client.get(f"/api/roster/semester-export/render-missing/{job_id}")
        assert_status(progress, 200)
        state = progress.json()
        if state["status"] != "running":
            return state
        time.sleep(0.2)
    raise AssertionError("補渲染 job 逾時未完成")


def test_project_completion_locks_content_and_supervisor_reopens(monkeypatch, tmp_path):
    use_tmp_uploads(monkeypatch, tmp_path)

    with started_client() as client:
        login(client)
        supervisor, supervisor_password = create_user(client, "supervisor")
        managed_teacher, teacher_password = create_user(client, "teacher")
        period = create_active_period(client)
        template_id, _ = create_template_with_page(client, period_id=period["id"])
        campus_id, classroom_id = create_scoped_classroom(client)
        set_campus_supervisor_scope(client, campus_id, supervisor["id"])
        set_classroom_teachers(client, classroom_id, [managed_teacher["id"]])
        overview = client.get("/api/organization/overview")
        work_slot_id = next(
            slot["id"]
            for slot in overview.json()["work_slots"]
            if slot["classroom_id"] == classroom_id
            and slot["template_period_id"] == period["id"]
        )
        empty_create = client.post(
            f"/api/organization/classrooms/{classroom_id}/projects",
            json={
                "name": unique_name("empty_blocked"),
                "template_id": template_id,
                "work_slot_id": work_slot_id,
            },
        )
        assert_status(empty_create, 409)
        student_name = unique_name("completion_student")
        add_classroom_members(client, classroom_id, [student_name])
        project_id = create_classroom_project(
            client,
            classroom_id,
            template_id,
        )

        # 老師完成後內容鎖定
        client.cookies.clear()
        login(client, managed_teacher["username"], teacher_password)

        students = add_students(client, project_id, [student_name])
        complete = client.post(f"/api/projects/{project_id}/complete")
        assert_status(complete, 200)
        assert complete.json()["completed_at"] is not None

        # 完成後：內容修改一律 403
        album_name = client.put(
            f"/api/projects/{project_id}/students/{students[student_name]}/album-name",
            json={"album_name": "小明"},
        )
        assert_status(album_name, 403)
        photo = client.post(
            revisioned_project_url(
                client,
                project_id,
                f"/api/projects/{project_id}/students/{students[student_name]}/pages/0/photos/1",
            ),
            files={"file": ("smoke.jpg", jpeg_bytes(), "image/jpeg")},
        )
        assert_status(photo, 403)
        texts = client.put(
            revisioned_project_url(client, project_id, f"/api/projects/{project_id}/label_texts"),
            json={"0": {"1": "改字"}},
        )
        assert_status(texts, 403)
        skip = client.patch(
            revisioned_project_url(
                client,
                project_id,
                f"/api/projects/{project_id}/students/{students[student_name]}/pages/0/skip",
            ),
            json={"skip": True},
        )
        assert_status(skip, 403)

        # 渲染與下載不算內容修改，仍可執行
        render = client.post(
            f"/api/projects/{project_id}/students/{students[student_name]}/render"
        )
        assert_status(render, 200)

        # 老師自己不能退回
        teacher_reopen = client.post(f"/api/projects/{project_id}/reopen")
        assert_status(teacher_reopen, 403)

        # 管轄主管視角：老師進度看得到 completed_at；退回後恢復可編輯
        client.cookies.clear()
        login(client, supervisor["username"], supervisor_password)
        semester_id = reporting_semester_id(client, [period["id"]])
        progress = client.get(
            "/api/roster/teacher-progress",
            params={"semester_id": semester_id},
        )
        assert_status(progress, 200)
        teacher_projects = {
            project["project_id"]: project
            for classroom in progress.json()["classrooms"]
            for slot in classroom["slots"]
            for project in slot["projects"]
        }
        assert teacher_projects[project_id]["completed_at"] is not None
        assert teacher_projects[project_id]["workflow_status"] == "submitted_locked"

        reopen = client.post(f"/api/projects/{project_id}/reopen")
        assert_status(reopen, 200)

        client.cookies.clear()
        login(client, managed_teacher["username"], teacher_password)
        album_name_after_reopen = client.put(
            f"/api/projects/{project_id}/students/{students[student_name]}/album-name",
            json={"album_name": "小明"},
        )
        assert_status(album_name_after_reopen, 409)
        assert (
            album_name_after_reopen.json()["detail"]["code"]
            == "roster_album_name_authority"
        )
        skip_after_reopen = client.patch(
            revisioned_project_url(
                client,
                project_id,
                f"/api/projects/{project_id}/students/"
                f"{students[student_name]}/pages/0/skip",
            ),
            json={"skip": True},
        )
        assert_status(skip_after_reopen, 200)


def _project_students_by_id(client: TestClient, project_id: int) -> dict[int, dict]:
    detail = client.get(f"/api/projects/{project_id}")
    assert_status(detail, 200)
    return {student["id"]: student for student in detail.json()["students"]}


def test_student_completion_locks_downloads_and_reopen_levels(monkeypatch, tmp_path):
    use_tmp_uploads(monkeypatch, tmp_path)

    with started_client() as client:
        login(client)
        supervisor, supervisor_password = create_user(client, "supervisor")
        teacher, teacher_password = create_user(client, "teacher")
        period = create_active_period(client)
        template_id, _ = create_template_with_page(client, period_id=period["id"])
        campus_id, classroom_id = create_scoped_classroom(client)
        set_campus_supervisor_scope(client, campus_id, supervisor["id"])
        set_classroom_teachers(client, classroom_id, [teacher["id"]])
        name_a = unique_name("完成甲")
        name_b = unique_name("進行乙")
        add_classroom_members(client, classroom_id, [name_a, name_b])
        project_id = create_classroom_project(client, classroom_id, template_id)

        client.cookies.clear()
        login(client, teacher["username"], teacher_password)
        students = add_students(client, project_id, [name_a, name_b])
        student_a = students[name_a]
        student_b = students[name_b]

        # 內容未填滿不可標記完成：409 附照片/文字計數
        incomplete_a = client.post(
            f"/api/projects/{project_id}/students/{student_a}/complete"
        )
        assert_status(incomplete_a, 409)
        incomplete_detail = incomplete_a.json()["detail"]
        assert incomplete_detail["code"] == "student_content_incomplete"
        assert incomplete_detail["photo_filled"] == 0
        assert incomplete_detail["photo_total"] == 1
        assert incomplete_detail["text_filled"] == 0
        assert incomplete_detail["text_total"] == 1

        # 以全班文字補滿文字格（計入已填）、補上 A 的照片後才可標記
        class_fill = client.put(
            revisioned_project_url(
                client, project_id, f"/api/projects/{project_id}/label_texts"
            ),
            json={"0": {"1": "完成前全班文字"}},
        )
        assert_status(class_fill, 200)
        photo_only_incomplete = client.post(
            f"/api/projects/{project_id}/students/{student_a}/complete"
        )
        assert_status(photo_only_incomplete, 409)
        assert (
            photo_only_incomplete.json()["detail"]["photo_filled"] == 0
        )
        photo_fill_a = client.post(
            revisioned_project_url(
                client,
                project_id,
                f"/api/projects/{project_id}/students/{student_a}/pages/0/photos/1",
            ),
            files={"file": ("fill.jpg", jpeg_bytes(), "image/jpeg")},
        )
        assert_status(photo_fill_a, 200)

        # 標記單生完成：冪等回傳既有時間戳，全班未成立
        complete_a = client.post(
            f"/api/projects/{project_id}/students/{student_a}/complete"
        )
        assert_status(complete_a, 200)
        first_completed_at = complete_a.json()["completed_at"]
        assert first_completed_at is not None
        assert complete_a.json()["project_completed_at"] is None
        complete_a_again = client.post(
            f"/api/projects/{project_id}/students/{student_a}/complete"
        )
        assert_status(complete_a_again, 200)
        assert complete_a_again.json()["completed_at"] == first_completed_at

        detail_students = _project_students_by_id(client, project_id)
        assert detail_students[student_a]["completed_at"] is not None
        assert detail_students[student_b]["completed_at"] is None

        # A 的逐學生寫入全部 409；B 完全不受影響
        photo_a = client.post(
            revisioned_project_url(
                client,
                project_id,
                f"/api/projects/{project_id}/students/{student_a}/pages/0/photos/1",
            ),
            files={"file": ("locked.jpg", jpeg_bytes(), "image/jpeg")},
        )
        assert_status(photo_a, 409)
        assert photo_a.json()["detail"]["code"] == "student_completed_locked"
        texts_a = client.put(
            revisioned_project_url(
                client,
                project_id,
                f"/api/projects/{project_id}/students/{student_a}/pages/0/texts",
            ),
            json={"1": "甲的個人文字"},
        )
        assert_status(texts_a, 409)
        assert texts_a.json()["detail"]["code"] == "student_completed_locked"
        skip_a = client.patch(
            revisioned_project_url(
                client,
                project_id,
                f"/api/projects/{project_id}/students/{student_a}/pages/0/skip",
            ),
            json={"skip": True},
        )
        assert_status(skip_a, 409)
        assert skip_a.json()["detail"]["code"] == "student_completed_locked"
        texts_b = client.put(
            revisioned_project_url(
                client,
                project_id,
                f"/api/projects/{project_id}/students/{student_b}/pages/0/texts",
            ),
            json={"1": "乙的個人文字"},
        )
        assert_status(texts_b, 200)

        # 任一學生完成後全班文字整份硬鎖
        class_texts = client.put(
            revisioned_project_url(
                client, project_id, f"/api/projects/{project_id}/label_texts"
            ),
            json={"0": {"1": "全班文字"}},
        )
        assert_status(class_texts, 409)
        class_texts_detail = class_texts.json()["detail"]
        assert class_texts_detail["code"] == "class_texts_locked_by_completed_students"
        assert class_texts_detail["completed_student_names"] == [name_a]

        # 批次文字含已完成學生 → 整批 422，B 也不寫入
        batch_texts = client.put(
            revisioned_project_url(
                client, project_id, f"/api/projects/{project_id}/batch/texts"
            ),
            json={
                "students": {
                    str(student_a): {"0": {"1": "批次甲"}},
                    str(student_b): {"0": {"1": "批次乙"}},
                }
            },
        )
        assert_status(batch_texts, 422)
        assert (
            batch_texts.json()["detail"]["code"]
            == "batch_texts_contains_completed_students"
        )
        detail_students = _project_students_by_id(client, project_id)
        assert (
            detail_students[student_b]["pages_data"][0]["label_texts"]["1"]
            == "乙的個人文字"
        )

        # 共用照片套用全班：自動排除已完成學生且回應揭露
        detail_students = _project_students_by_id(client, project_id)
        photo_a_before_shared = (
            detail_students[student_a]["pages_data"][0]["photos"]["1"]
        )
        shared_all = client.post(
            revisioned_project_url(
                client,
                project_id,
                f"/api/projects/{project_id}/photos/shared/pages/0/slots/1",
            ),
            files={"file": ("shared.jpg", jpeg_bytes(), "image/jpeg")},
        )
        assert_status(shared_all, 200)
        assert shared_all.json()["updated"] == 1
        assert shared_all.json()["skipped_completed_student_ids"] == [student_a]
        detail_students = _project_students_by_id(client, project_id)
        # A 已完成：既有照片不被共用上傳覆蓋
        assert (
            detail_students[student_a]["pages_data"][0]["photos"]["1"]
            == photo_a_before_shared
        )
        photo_b_shared = detail_students[student_b]["pages_data"][0]["photos"]["1"]
        assert photo_b_shared and photo_b_shared != photo_a_before_shared

        # 共用照片指定名單含已完成學生 → 422
        shared_targeted = client.post(
            revisioned_project_url(
                client,
                project_id,
                f"/api/projects/{project_id}/photos/shared/pages/0/slots/1",
            ),
            files={"file": ("shared2.jpg", jpeg_bytes(), "image/jpeg")},
            data={"student_ids": f"[{student_a}]"},
        )
        assert_status(shared_targeted, 422)
        assert (
            shared_targeted.json()["detail"]["code"]
            == "shared_photo_targets_completed_students"
        )

        # 下載閘門：A 可下載、B 409、全班 ZIP 409 且附 n/N
        render_a = client.post(
            f"/api/projects/{project_id}/students/{student_a}/render"
        )
        assert_status(render_a, 200)
        download_a = client.get(
            f"/api/projects/{project_id}/students/{student_a}/pdf"
        )
        assert_status(download_a, 200)
        assert download_a.content.startswith(b"%PDF")
        download_b = client.get(
            f"/api/projects/{project_id}/students/{student_b}/pdf"
        )
        assert_status(download_b, 409)
        assert download_b.json()["detail"]["code"] == "student_not_completed"
        class_zip = client.get(f"/api/projects/{project_id}/download/all")
        assert_status(class_zip, 409)
        class_zip_detail = class_zip.json()["detail"]
        assert class_zip_detail["code"] == "class_zip_requires_full_completion"
        assert class_zip_detail["completed"] == 1
        assert class_zip_detail["total"] == 2

        # 老師退單生一律 403；主管退回後恢復可編輯
        teacher_reopen_a = client.post(
            f"/api/projects/{project_id}/students/{student_a}/reopen"
        )
        assert_status(teacher_reopen_a, 403)
        client.cookies.clear()
        login(client, supervisor["username"], supervisor_password)
        supervisor_reopen_a = client.post(
            f"/api/projects/{project_id}/students/{student_a}/reopen"
        )
        assert_status(supervisor_reopen_a, 200)
        client.cookies.clear()
        login(client, teacher["username"], teacher_password)
        skip_a_after_reopen = client.patch(
            revisioned_project_url(
                client,
                project_id,
                f"/api/projects/{project_id}/students/{student_a}/pages/0/skip",
            ),
            json={"skip": False},
        )
        assert_status(skip_a_after_reopen, 200)

        # 最後一位學生完成 → 同 transaction 自動成立全班完成
        assert_status(
            client.post(f"/api/projects/{project_id}/students/{student_a}/complete"),
            200,
        )
        complete_b = client.post(
            f"/api/projects/{project_id}/students/{student_b}/complete"
        )
        assert_status(complete_b, 200)
        assert complete_b.json()["project_completed_at"] is not None
        full_detail = client.get(f"/api/projects/{project_id}")
        assert_status(full_detail, 200)
        assert full_detail.json()["completed_at"] is not None

        # 改名不受完成狀態限制：全部學生輸出清空、完成狀態不變
        renamed_name = unique_name("完成後改名")
        rename = client.patch(
            f"/api/projects/{project_id}", data={"name": renamed_name}
        )
        assert_status(rename, 200)
        renamed_detail = client.get(f"/api/projects/{project_id}")
        assert_status(renamed_detail, 200)
        assert renamed_detail.json()["name"] == renamed_name
        assert renamed_detail.json()["completed_at"] is not None
        assert all(
            student["output_filename"] is None
            for student in renamed_detail.json()["students"]
        )

        # 全班完成成立後主管退單生：project.completed_at 一併清除，其他學生保留
        client.cookies.clear()
        login(client, supervisor["username"], supervisor_password)
        assert_status(
            client.post(f"/api/projects/{project_id}/students/{student_a}/reopen"),
            200,
        )
        partial_detail = client.get(f"/api/projects/{project_id}")
        assert_status(partial_detail, 200)
        assert partial_detail.json()["completed_at"] is None
        partial_students = {
            student["id"]: student for student in partial_detail.json()["students"]
        }
        assert partial_students[student_a]["completed_at"] is None
        assert partial_students[student_b]["completed_at"] is not None

        # 全班退回：全部學生 completed_at 清空，無殘留個別鎖
        assert_status(client.post(f"/api/projects/{project_id}/reopen"), 200)
        reopened_students = _project_students_by_id(client, project_id)
        assert all(
            student["completed_at"] is None
            for student in reopened_students.values()
        )

        # 手動全班完成不回填學生時間戳；單生下載閘門仍以 predicate 放行
        client.cookies.clear()
        login(client, teacher["username"], teacher_password)
        assert_status(client.post(f"/api/projects/{project_id}/complete"), 200)
        manual_students = _project_students_by_id(client, project_id)
        assert all(
            student["completed_at"] is None
            for student in manual_students.values()
        )
        # 輸出已因改名清空：閘門放行後下載端點就地補渲最新內容（不再 404）
        manual_download_a = client.get(
            f"/api/projects/{project_id}/students/{student_a}/pdf"
        )
        assert_status(manual_download_a, 200)
        assert manual_download_a.content.startswith(b"%PDF")


def test_render_missing_fills_absent_pdfs(monkeypatch, tmp_path):
    use_tmp_uploads(monkeypatch, tmp_path)

    with started_client() as client:
        login(client)
        period = create_active_period(client)
        project = create_period_template_project(
            client,
            period["id"],
            ["王小明", "李小華"],
        )
        students = add_students(client, project, ["王小明", "李小華"])
        # 只給王小明照片；兩人都未渲染
        photo = client.post(
            revisioned_project_url(
                client,
                project,
                f"/api/projects/{project}/students/{students['王小明']}/pages/0/photos/1",
            ),
            files={"file": ("smoke.jpg", jpeg_bytes(), "image/jpeg")},
        )
        assert_status(photo, 200)

        before = get_semester_preview(client, [period["id"]])
        assert all(
            not entry["has_pdf"]
            for child in preview_children(before.json())
            for entry in child_entries(child)
        )

        result = start_and_wait_render_job(client, [period["id"]])
        assert result["status"] == "done"
        assert result["total"] == 2
        assert result["done"] == 2
        assert result["rendered"] == 2
        assert result["errors"] == []

        after = get_semester_preview(client, [period["id"]])
        assert all(
            entry["has_pdf"]
            for child in preview_children(after.json())
            for entry in child_entries(child)
        )

        # 再跑一次：已全數渲染，rendered=0（冪等）
        rerun = start_and_wait_render_job(client, [period["id"]])
        assert rerun["status"] == "done"
        assert rerun["rendered"] == 0

        # 不存在的 job 回 404
        missing_job = client.get("/api/roster/semester-export/render-missing/nonexistent")
        assert_status(missing_job, 404)


def test_render_missing_keeps_partial_success_and_progress_after_one_failure(
    monkeypatch,
    tmp_path,
):
    use_tmp_uploads(monkeypatch, tmp_path)
    with started_client() as client:
        login(client)
        period = create_active_period(client)
        project_id = create_period_template_project(
            client,
            period["id"],
            ["第一位失敗", "第二位成功"],
        )
        student_ids = add_students(client, project_id, ["第一位失敗", "第二位成功"])
        semester_id = reporting_semester_id(client, [period["id"]])

    rendered_students = []

    def render_one(project, student, requested_project_id, db):
        rendered_students.append(student.name)
        if student.id == student_ids["第一位失敗"]:
            raise RuntimeError("simulated single render failure")
        return {"pdf": "unused", "pages": 1}

    monkeypatch.setattr(
        semester_render_service,
        "render_and_save_student_album",
        render_one,
    )
    progress = []
    db = SessionLocal()
    try:
        result = semester_render_service.render_missing_semester_albums(
            db,
            semester_id,
            [period["id"]],
            progress_callback=lambda done, total: progress.append((done, total)),
        )
    finally:
        db.close()

    assert rendered_students == ["第一位失敗", "第二位成功"]
    assert result["rendered"] == 1
    assert len(result["errors"]) == 1
    assert result["errors"][0]["student"] == "第一位失敗"
    assert result["errors"][0]["project"]
    assert result["errors"][0]["error"] == "產生失敗"
    assert progress == [(0, 2), (1, 2), (2, 2)]


def test_semester_export_zip_structure(monkeypatch, tmp_path):
    use_tmp_uploads(monkeypatch, tmp_path)

    with started_client() as client:
        login(client)
        teacher, _ = create_user(client, "teacher")
        period_a = create_active_period(client)
        period_b = create_active_period(client)
        template_a_id, _ = create_template_with_page(client, period_id=period_a["id"])
        template_b_id, _ = create_template_with_page(client, period_id=period_b["id"])
        _, classroom_id = create_scoped_classroom(client)
        set_classroom_teachers(client, classroom_id, [teacher["id"]])
        add_classroom_members(client, classroom_id, ["王小明", "李小華"])
        project_a = create_classroom_project(client, classroom_id, template_a_id)
        project_b = create_classroom_project(client, classroom_id, template_b_id)

        students_a = add_students(client, project_a, ["王小明", "李小華"])
        students_b = add_students(client, project_b, ["王小明"])

        # 王小明兩期都渲染；李小華不渲染（應出現在匯出說明）
        for project_id, student_id in (
            (project_a, students_a["王小明"]),
            (project_b, students_b["王小明"]),
        ):
            photo = client.post(
                revisioned_project_url(
                    client,
                    project_id,
                    f"/api/projects/{project_id}/students/{student_id}/pages/0/photos/1",
                ),
                files={"file": ("smoke.jpg", jpeg_bytes(), "image/jpeg")},
            )
            assert_status(photo, 200)
            render = client.post(f"/api/projects/{project_id}/students/{student_id}/render")
            assert_status(render, 200)

        # 老師手動刪除王小明期1 的第 1 頁 → 匯出說明應標註缺頁
        skip_response = client.patch(
            revisioned_project_url(
                client,
                project_a,
                f"/api/projects/{project_a}/students/{students_a['王小明']}/pages/0/skip",
            ),
            json={"skip": True},
        )
        assert_status(skip_response, 200)
        semester_id = reporting_semester_id(
            client,
            [period_a["id"], period_b["id"]],
        )

        download = client.get(
            "/api/roster/semester-export/download",
            params={
                "semester_id": semester_id,
                "period_ids": [period_a["id"], period_b["id"]],
                "mode": "print",
            },
        )
        assert_status(download, 200)
        assert download.headers["content-type"] == "application/zip"

        with ZipFile(BytesIO(download.content)) as zip_archive:
            entry_names = zip_archive.namelist()
            manifest = zip_archive.read("匯出說明.txt").decode("utf-8")
        # 結構：校別/班級/孩子/期別_孩子.pdf；兩期以上另附全期合併
        ming_entries = {name for name in entry_names if "/王小明/" in name}
        assert {name.rsplit("/", 1)[-1] for name in ming_entries} == {
            f"{period_a['name']}_王小明.pdf",
            f"{period_b['name']}_王小明.pdf",
            "全期合併_王小明.pdf",
        }
        # 未渲染的李小華不進 ZIP，但列在匯出說明；班級對照含兩位孩子
        assert not any("/李小華/" in name for name in entry_names)
        assert "【班級對照】" in manifest
        assert "李小華" in manifest and "尚未產生 PDF" in manifest
        # 缺頁備註：王小明期1 老師刪除了第 1 頁
        assert "缺頁備註" in manifest
        assert "老師刪除了第 1 頁" in manifest

        # roster_child_ids 篩選：只勾李小華 → ZIP 不含王小明
        hua_child_id = roster_child_id_of(students_a["李小華"])
        filtered = client.get(
            "/api/roster/semester-export/download",
            params={
                "semester_id": semester_id,
                "period_ids": [period_a["id"], period_b["id"]],
                "mode": "print",
                "roster_child_ids": [hua_child_id],
            },
        )
        assert_status(filtered, 200)
        with ZipFile(BytesIO(filtered.content)) as zip_archive:
            filtered_names = zip_archive.namelist()
            filtered_manifest = zip_archive.read("匯出說明.txt").decode("utf-8")
        assert not any("/王小明/" in name for name in filtered_names)
        assert "王小明" not in filtered_manifest
        assert "李小華" in filtered_manifest


def test_semester_export_zip_paths_cannot_escape_with_database_names(monkeypatch):
    first_pdf_key = get_student_pdf_key(9, 101)
    second_pdf_key = get_student_pdf_key(10, 102)
    preview = {
        "periods": [
            {
                "id": 1,
                "template_period_id": 1,
                "name": "C:\\期別/上學期",
                "department": "infant",
            }
        ],
        "classroom_groups": [
            {
                "campus_name": ".",
                "classroom_name": "/根目錄",
                "children": [
                    {
                        "roster_child_id": 1,
                        "name": ".",
                        "latest_classroom": {
                            "campus_name": ".",
                            "classroom_name": "/根目錄",
                        },
                        "cells": [{
                            "template_period_id": 1,
                            "status": "ready",
                            "entries": [{
                                "project_id": 9,
                                "project_name": "/根目錄",
                                "student_id": 101,
                                "campus_name": ".",
                                "classroom_name": "/根目錄",
                                "skipped_pages": [],
                            }],
                        }],
                    },
                ],
            },
            {
                "campus_name": ".",
                "classroom_name": "\\\\server\\share",
                "children": [
                    {
                        "roster_child_id": 2,
                        "name": "..",
                        "latest_classroom": {
                            "campus_name": ".",
                            "classroom_name": "\\\\server\\share",
                        },
                        "cells": [{
                            "template_period_id": 1,
                            "status": "ready",
                            "entries": [{
                                "project_id": 10,
                                "project_name": "\\\\server\\share",
                                "student_id": 102,
                                "campus_name": ".",
                                "classroom_name": "\\\\server\\share",
                                "skipped_pages": [],
                            }],
                        }],
                    },
                ],
            },
        ],
        "unlinked": [],
    }
    projects = [
        SimpleNamespace(
            id=9,
            students=[
                SimpleNamespace(id=101, output_filename=first_pdf_key),
            ],
        ),
        SimpleNamespace(
            id=10,
            students=[
                SimpleNamespace(id=102, output_filename=second_pdf_key),
            ],
        ),
    ]
    monkeypatch.setattr(
        semester_export_service,
        "build_semester_export_preview",
        lambda db, semester_id, period_ids: preview,
    )
    monkeypatch.setattr(
        semester_export_service,
        "load_export_projects",
        lambda db, semester_id, period_ids: projects,
    )
    monkeypatch.setattr(semester_export_service, "get_storage", lambda: object())
    monkeypatch.setattr(
        semester_export_service,
        "load_output_keys_by_project",
        lambda storage, loaded_projects: {
            9: {first_pdf_key},
            10: {second_pdf_key},
        },
    )

    child_plans, _ = semester_export_service._plan_semester_export_zip(
        None,
        7,
        [1],
        "print",
    )

    assert child_plans == [
        {
            "files": [
                ("unnamed/_根目錄/unnamed/C__期別_上學期_unnamed.pdf", first_pdf_key),
            ],
            "merged_path": None,
        },
        {
            "files": [
                (
                    "unnamed/__server_share/unnamed/C__期別_上學期_unnamed.pdf",
                    second_pdf_key,
                ),
            ],
            "merged_path": None,
        },
    ]
    for plan in child_plans:
        for archive_path, _ in plan["files"]:
            assert not archive_path.startswith(("/", "\\"))
            assert ":" not in archive_path
            assert "\\" not in archive_path
            assert all(
                segment not in {".", ".."}
                for segment in archive_path.split("/")
            )
