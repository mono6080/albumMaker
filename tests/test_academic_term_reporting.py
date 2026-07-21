"""正式學期報表的歷史、狀態與匯出安全契約。"""

from io import BytesIO
from zipfile import ZipFile

from openpyxl import load_workbook
from pypdf import PdfReader

from database import (
    AcademicTerm,
    AcademicTermClassroom,
    Classroom,
    Project,
    RosterChild,
    SessionLocal,
    Student,
)
from routers import roster as roster_router
from services import semester_render_service
from tests.helpers import (
    USER_PASSWORD,
    assert_status,
    create_template_with_page,
    create_user,
    login,
    started_client,
    unique_name,
    use_tmp_uploads,
)
from tests.test_roster import (
    add_classroom_members,
    add_students,
    create_active_period,
    create_classroom_project,
    create_scoped_classroom,
    get_semester_preview,
    preview_children,
    reporting_term_id,
    set_campus_supervisor_scope,
    set_classroom_teachers,
)


def _project_progress(payload: dict, project_id: int) -> dict:
    return next(
        project
        for classroom in payload["classrooms"]
        for slot in classroom["slots"]
        for project in slot["projects"]
        if project["project_id"] == project_id
    )


def _project_slot(payload: dict, project_id: int) -> dict:
    return next(
        slot
        for classroom in payload["classrooms"]
        for slot in classroom["slots"]
        if any(
            project["project_id"] == project_id
            for project in slot["projects"]
        )
    )


def test_progress_keeps_started_slot_archived_and_separates_status_axes():
    with started_client() as client:
        login(client)
        teacher, _ = create_user(client, "teacher")
        period = create_active_period(client)
        template_id, _ = create_template_with_page(client, period_id=period["id"])
        _, classroom_id = create_scoped_classroom(client)
        set_classroom_teachers(client, classroom_id, [teacher["id"]])
        add_classroom_members(
            client,
            classroom_id,
            [unique_name("missing_photo_a"), unique_name("missing_photo_b")],
        )
        project_id = create_classroom_project(client, classroom_id, template_id)
        academic_term_id = reporting_term_id(client, [period["id"]])

        progress = client.get(
            "/api/roster/teacher-progress",
            params={"academic_term_id": academic_term_id},
        )
        assert_status(progress, 200)
        project = _project_progress(progress.json(), project_id)
        assert project["content_status"] == "incomplete"
        assert project["workflow_status"] == "working"
        assert "export_status" not in project
        assert project["attention_codes"] == []

        complete = client.post(f"/api/projects/{project_id}/complete")
        assert_status(complete, 200)
        submitted_progress = client.get(
            "/api/roster/teacher-progress",
            params={"academic_term_id": academic_term_id},
        )
        submitted_project = _project_progress(
            submitted_progress.json(),
            project_id,
        )
        assert submitted_project["workflow_status"] == "submitted_locked"
        assert submitted_project["content_status"] == "incomplete"
        assert submitted_project["attention_codes"] == [
            "submitted_with_missing_photos",
            "submitted_with_missing_texts",
        ]

        archived = client.delete(f"/api/projects/{project_id}")
        assert_status(archived, 200)
        archived_progress = client.get(
            "/api/roster/teacher-progress",
            params={"academic_term_id": academic_term_id},
        )
        assert_status(archived_progress, 200)
        classroom = next(
            classroom
            for classroom in archived_progress.json()["classrooms"]
            if classroom["classroom_id"] == classroom_id
        )
        slot = next(
            slot
            for slot in classroom["slots"]
            if slot["template_period_id"] == period["id"]
        )
        assert slot["creation_status"] == "archived"
        assert slot["started_at"] is not None
        assert slot["projects"] == []


def test_closed_term_does_not_seed_children_from_current_roster():
    with started_client() as client:
        login(client)
        teacher, _ = create_user(client, "teacher")
        period = create_active_period(client)
        template_id, _ = create_template_with_page(client, period_id=period["id"])
        _, classroom_id = create_scoped_classroom(client)
        set_classroom_teachers(client, classroom_id, [teacher["id"]])
        original_name = unique_name("closed_term_original")
        add_classroom_members(client, classroom_id, [original_name])
        create_classroom_project(client, classroom_id, template_id)
        academic_term_id = reporting_term_id(client, [period["id"]])

        db = SessionLocal()
        try:
            term = db.get(AcademicTerm, academic_term_id)
            original_status = term.status
            term.status = "closed"
            db.commit()
        finally:
            db.close()

        current_only_name = unique_name("current_roster_only")
        try:
            add_classroom_members(client, classroom_id, [current_only_name])
            preview = client.get(
                "/api/roster/semester-export",
                params={
                    "academic_term_id": academic_term_id,
                    "period_ids": [period["id"]],
                },
            )
            assert_status(preview, 200)
            child_names = {
                child["name"] for child in preview_children(preview.json())
            }
            assert original_name in child_names
            assert current_only_name not in child_names
        finally:
            db = SessionLocal()
            try:
                db.get(AcademicTerm, academic_term_id).status = original_status
                db.commit()
            finally:
                db.close()


def test_closed_term_keeps_snapshot_student_without_any_project():
    with started_client() as client:
        login(client)
        period = create_active_period(client)
        _, classroom_id = create_scoped_classroom(client)
        student_name = unique_name("closed_no_project")
        add_classroom_members(client, classroom_id, [student_name])
        academic_term_id = reporting_term_id(client, [period["id"]])

        db = SessionLocal()
        try:
            term = db.get(AcademicTerm, academic_term_id)
            original_status = term.status
            term.status = "closed"
            db.commit()
        finally:
            db.close()

        try:
            preview = client.get(
                "/api/roster/semester-export",
                params={
                    "academic_term_id": academic_term_id,
                    "period_ids": [period["id"]],
                },
            )
            assert_status(preview, 200)
            child = next(
                child
                for child in preview_children(preview.json())
                if child["name"] == student_name
            )
            assert [cell["status"] for cell in child["cells"]] == ["no_album"]
            assert child["cells"][0]["entries"] == []
        finally:
            db = SessionLocal()
            try:
                db.get(AcademicTerm, academic_term_id).status = original_status
                db.commit()
            finally:
                db.close()


def test_teacher_overview_excel_escapes_formula_like_user_text():
    with started_client() as client:
        login(client)
        teacher_response = client.post(
            "/api/users/",
            json={
                "username": unique_name("formula_teacher"),
                "display_name": "=1+1",
                "password": USER_PASSWORD,
                "role": "teacher",
            },
        )
        assert_status(teacher_response, 201)
        teacher_id = teacher_response.json()["id"]
        period_name = unique_name("=period")
        period_response = client.post(
            "/api/templates/periods",
            data={
                "name": period_name,
                "department": "infant",
                "status": "active",
            },
        )
        assert_status(period_response, 200)
        period = period_response.json()
        template_id, _ = create_template_with_page(client, period_id=period["id"])

        campus_name = unique_name(" +campus")
        campus_response = client.post(
            "/api/organization/campuses",
            json={"name": campus_name},
        )
        assert_status(campus_response, 201)
        campus_id = campus_response.json()["id"]
        classroom_name = unique_name("-classroom")
        classroom_response = client.post(
            "/api/organization/classrooms",
            json={
                "campus_id": campus_id,
                "department": "infant",
                "name": classroom_name,
            },
        )
        assert_status(classroom_response, 201)
        classroom_id = classroom_response.json()["id"]
        set_classroom_teachers(client, classroom_id, [teacher_id])
        student_name = unique_name("@student")
        add_classroom_members(client, classroom_id, [student_name])

        overview = client.get("/api/organization/overview")
        work_slot_id = next(
            slot["id"]
            for slot in overview.json()["work_slots"]
            if slot["classroom_id"] == classroom_id
            and slot["template_period_id"] == period["id"]
        )
        project_name = unique_name("=project")
        project_response = client.post(
            f"/api/organization/classrooms/{classroom_id}/projects",
            json={
                "name": project_name,
                "template_id": template_id,
                "work_slot_id": work_slot_id,
            },
        )
        assert_status(project_response, 201)
        academic_term_id = reporting_term_id(client, [period["id"]])

        export = client.get(
            "/api/roster/teacher-overview/export",
            params={
                "academic_term_id": academic_term_id,
                "campus_id": campus_id,
                "classroom_id": classroom_id,
            },
        )
        assert_status(export, 200)
        workbook = load_workbook(BytesIO(export.content), data_only=False)
        slot_row = next(
            row
            for row in workbook["班級期別"].iter_rows(min_row=2)
            if row[7].value == 1
        )
        student_row = next(workbook["學生明細"].iter_rows(min_row=2))
        formula_like_cells = [
            slot_row[0],
            slot_row[2],
            slot_row[3],
            slot_row[4],
            student_row[5],
            student_row[6],
            student_row[7],
        ]
        assert all(cell.data_type != "f" for cell in formula_like_cells)
        assert all(str(cell.value).startswith("'") for cell in formula_like_cells)


def test_render_request_validates_term_period_before_starting_job(monkeypatch):
    started_jobs = []
    monkeypatch.setattr(
        roster_router,
        "start_render_missing_job",
        lambda *args: started_jobs.append(args),
    )
    with started_client() as client:
        login(client)
        period = create_active_period(client)
        academic_term_id = reporting_term_id(client, [period["id"]])
        response = client.post(
            "/api/roster/semester-export/render-missing",
            json={
                "academic_term_id": academic_term_id,
                "period_ids": [2_000_000_000],
            },
        )
        assert_status(response, 422)
        assert response.json()["detail"]["code"] == "period_not_in_academic_term"
        assert started_jobs == []


def test_report_direct_term_id_hides_out_of_scope_and_nonreporting_terms():
    with started_client() as client:
        login(client)
        supervisor, supervisor_password = create_user(client, "supervisor")
        scoped_campus_id, _ = create_scoped_classroom(client)
        set_campus_supervisor_scope(
            client,
            scoped_campus_id,
            supervisor["id"],
            department="infant",
        )
        _, outside_classroom_id = create_scoped_classroom(client)

        db = SessionLocal()
        try:
            outside_classroom = db.get(Classroom, outside_classroom_id)
            closed_term = AcademicTerm(
                label=unique_name("outside_closed_term"),
                status="closed",
                created_by_name_snapshot="admin",
            )
            cancelled_term = AcademicTerm(
                label=unique_name("cancelled_term"),
                status="cancelled",
                created_by_name_snapshot="admin",
            )
            db.add_all([closed_term, cancelled_term])
            db.flush()
            db.add_all([
                AcademicTermClassroom(
                    academic_term_id=closed_term.id,
                    classroom_id=outside_classroom.id,
                    campus_id_snapshot=outside_classroom.campus_id,
                    campus_name_snapshot=outside_classroom.campus.name,
                    classroom_name_snapshot=outside_classroom.name,
                    department=outside_classroom.department,
                ),
                AcademicTermClassroom(
                    academic_term_id=cancelled_term.id,
                    classroom_id=outside_classroom.id,
                    campus_id_snapshot=outside_classroom.campus_id,
                    campus_name_snapshot=outside_classroom.campus.name,
                    classroom_name_snapshot=outside_classroom.name,
                    department=outside_classroom.department,
                ),
            ])
            db.commit()
            closed_term_id = closed_term.id
            cancelled_term_id = cancelled_term.id
        finally:
            db.close()

        client.cookies.clear()
        login(client, supervisor["username"], supervisor_password)
        hidden_teacher = client.get(
            "/api/roster/teacher-progress",
            params={"academic_term_id": closed_term_id},
        )
        assert_status(hidden_teacher, 404)
        hidden_semester = client.get(
            "/api/roster/semester-export",
            params={
                "academic_term_id": closed_term_id,
                "period_ids": [2_000_000_000],
            },
        )
        assert_status(hidden_semester, 404)

        client.cookies.clear()
        login(client)
        cancelled_teacher = client.get(
            "/api/roster/teacher-progress",
            params={"academic_term_id": cancelled_term_id},
        )
        assert_status(cancelled_teacher, 404)
        cancelled_semester = client.get(
            "/api/roster/semester-export",
            params={
                "academic_term_id": cancelled_term_id,
                "period_ids": [2_000_000_000],
            },
        )
        assert_status(cancelled_semester, 404)


def test_reporting_term_list_only_exposes_supervisor_departments():
    with started_client() as client:
        login(client)
        supervisor, supervisor_password = create_user(client, "supervisor")
        infant_period = create_active_period(client, department="infant")
        academy_period = create_active_period(client, department="academy")
        campus_id, _ = create_scoped_classroom(client, department="infant")
        create_scoped_classroom(
            client,
            campus_id=campus_id,
            department="academy",
        )
        set_campus_supervisor_scope(
            client,
            campus_id,
            supervisor["id"],
            department="academy",
        )
        academic_term_id = reporting_term_id(
            client,
            [infant_period["id"], academy_period["id"]],
        )

        client.cookies.clear()
        login(client, supervisor["username"], supervisor_password)
        response = client.get("/api/roster/academic-terms")
        assert_status(response, 200)
        term = next(
            term
            for term in response.json()["terms"]
            if term["id"] == academic_term_id
        )
        assert academy_period["id"] in {
            period["template_period_id"] for period in term["periods"]
        }
        assert infant_period["id"] not in {
            period["template_period_id"] for period in term["periods"]
        }
        assert {
            period["department"] for period in term["periods"]
        } == {"academy"}


def test_semester_preview_hides_out_of_scope_period_ids_from_supervisor():
    with started_client() as client:
        login(client)
        supervisor, supervisor_password = create_user(client, "supervisor")
        infant_period = create_active_period(client, department="infant")
        academy_period = create_active_period(client, department="academy")
        campus_id, _ = create_scoped_classroom(client, department="infant")
        create_scoped_classroom(
            client,
            campus_id=campus_id,
            department="academy",
        )
        set_campus_supervisor_scope(
            client,
            campus_id,
            supervisor["id"],
            department="academy",
        )
        academic_term_id = reporting_term_id(
            client,
            [infant_period["id"], academy_period["id"]],
        )

        client.cookies.clear()
        login(client, supervisor["username"], supervisor_password)
        visible_preview = client.get(
            "/api/roster/semester-export",
            params={
                "academic_term_id": academic_term_id,
                "period_ids": [academy_period["id"]],
            },
        )
        assert_status(visible_preview, 200)
        assert [
            period["template_period_id"]
            for period in visible_preview.json()["periods"]
        ] == [academy_period["id"]]

        for hidden_period_ids in (
            [infant_period["id"]],
            [academy_period["id"], infant_period["id"]],
            [2_000_000_000],
        ):
            hidden_preview = client.get(
                "/api/roster/semester-export",
                params={
                    "academic_term_id": academic_term_id,
                    "period_ids": hidden_period_ids,
                },
            )
            assert_status(hidden_preview, 404)
            assert hidden_preview.json() == {"detail": "找不到期別"}


def test_closed_term_uses_term_student_name_snapshot_for_preview_and_zip(
    monkeypatch,
    tmp_path,
):
    use_tmp_uploads(monkeypatch, tmp_path)
    with started_client() as client:
        login(client)
        teacher, _ = create_user(client, "teacher")
        period = create_active_period(client)
        template_id, _ = create_template_with_page(client, period_id=period["id"])
        _, classroom_id = create_scoped_classroom(client)
        set_classroom_teachers(client, classroom_id, [teacher["id"]])
        historical_name = unique_name("historical_name")
        current_name = unique_name("current_name")
        add_classroom_members(client, classroom_id, [historical_name])
        project_id = create_classroom_project(client, classroom_id, template_id)
        student_id = add_students(
            client,
            project_id,
            [historical_name],
        )[historical_name]
        rendered = client.post(
            f"/api/projects/{project_id}/students/{student_id}/render"
        )
        assert_status(rendered, 200)
        academic_term_id = reporting_term_id(client, [period["id"]])

        db = SessionLocal()
        try:
            term = db.get(AcademicTerm, academic_term_id)
            original_status = term.status
            term.status = "closed"
            student = db.get(Student, student_id)
            db.get(RosterChild, student.roster_child_id).name = current_name
            db.commit()
        finally:
            db.close()

        try:
            preview = client.get(
                "/api/roster/semester-export",
                params={
                    "academic_term_id": academic_term_id,
                    "period_ids": [period["id"]],
                },
            )
            assert_status(preview, 200)
            names = {child["name"] for child in preview_children(preview.json())}
            assert historical_name in names
            assert current_name not in names

            download = client.get(
                "/api/roster/semester-export/download",
                params={
                    "academic_term_id": academic_term_id,
                    "period_ids": [period["id"]],
                    "mode": "print",
                },
            )
            assert_status(download, 200)
            with ZipFile(BytesIO(download.content)) as zip_archive:
                entry_names = zip_archive.namelist()
            assert any(f"/{historical_name}/" in name for name in entry_names)
            assert not any(current_name in name for name in entry_names)
        finally:
            db = SessionLocal()
            try:
                db.get(AcademicTerm, academic_term_id).status = original_status
                db.commit()
            finally:
                db.close()


def test_semester_zip_uses_term_student_final_classroom_for_transferred_child(
    monkeypatch,
    tmp_path,
):
    use_tmp_uploads(monkeypatch, tmp_path)
    with started_client() as client:
        login(client)
        teacher, _ = create_user(client, "teacher")
        period = create_active_period(client)
        template_id, _ = create_template_with_page(client, period_id=period["id"])
        campus_id, source_classroom_id = create_scoped_classroom(client)
        _, target_classroom_id = create_scoped_classroom(
            client,
            campus_id=campus_id,
        )
        set_classroom_teachers(client, source_classroom_id, [teacher["id"]])
        student_name = unique_name("transferred_child")
        member_response = client.post(
            f"/api/organization/classrooms/{source_classroom_id}/members/batch",
            json={"members": [{"name": student_name}]},
        )
        assert_status(member_response, 201)
        member_id = member_response.json()["created"][0]["id"]
        project_id = create_classroom_project(
            client,
            source_classroom_id,
            template_id,
        )
        student_id = add_students(
            client,
            project_id,
            [student_name],
        )[student_name]
        rendered = client.post(
            f"/api/projects/{project_id}/students/{student_id}/render"
        )
        assert_status(rendered, 200)
        transferred = client.patch(
            f"/api/organization/classrooms/{source_classroom_id}/members/{member_id}",
            json={"target_classroom_id": target_classroom_id},
        )
        assert_status(transferred, 200)
        academic_term_id = reporting_term_id(client, [period["id"]])

        db = SessionLocal()
        try:
            term = db.get(AcademicTerm, academic_term_id)
            original_status = term.status
            term.status = "closed"
            campus_name = db.get(Classroom, target_classroom_id).campus.name
            source_classroom_name = db.get(Classroom, source_classroom_id).name
            target_classroom_name = db.get(Classroom, target_classroom_id).name
            db.commit()
        finally:
            db.close()

        try:
            preview = client.get(
                "/api/roster/semester-export",
                params={
                    "academic_term_id": academic_term_id,
                    "period_ids": [period["id"]],
                },
            )
            assert_status(preview, 200)
            child = next(
                child
                for child in preview_children(preview.json())
                if child["name"] == student_name
            )
            assert child["latest_classroom"]["classroom_id"] == target_classroom_id
            assert child["cells"][0]["entries"][0]["classroom_id"] == (
                source_classroom_id
            )

            download = client.get(
                "/api/roster/semester-export/download",
                params={
                    "academic_term_id": academic_term_id,
                    "period_ids": [period["id"]],
                    "mode": "print",
                },
            )
            assert_status(download, 200)
            with ZipFile(BytesIO(download.content)) as zip_archive:
                pdf_paths = [
                    name for name in zip_archive.namelist() if name.endswith(".pdf")
                ]
            assert len(pdf_paths) == 1
            assert pdf_paths[0].startswith(
                f"{campus_name}/{target_classroom_name}/{student_name}/"
            )
            assert f"/{source_classroom_name}/" not in pdf_paths[0]
        finally:
            db = SessionLocal()
            try:
                db.get(AcademicTerm, academic_term_id).status = original_status
                db.commit()
            finally:
                db.close()


def test_semester_zip_appends_merged_pdf_for_multi_period_child(
    monkeypatch,
    tmp_path,
):
    use_tmp_uploads(monkeypatch, tmp_path)
    with started_client() as client:
        login(client)
        teacher, _ = create_user(client, "teacher")
        period_a = create_active_period(client)
        period_b = create_active_period(client)
        template_a_id, _ = create_template_with_page(client, period_id=period_a["id"])
        template_b_id, _ = create_template_with_page(client, period_id=period_b["id"])
        _, classroom_id = create_scoped_classroom(client)
        set_classroom_teachers(client, classroom_id, [teacher["id"]])
        child_name = unique_name("merged_child")
        add_classroom_members(client, classroom_id, [child_name])
        for template_id in (template_a_id, template_b_id):
            project_id = create_classroom_project(client, classroom_id, template_id)
            student_id = add_students(client, project_id, [child_name])[child_name]
            rendered = client.post(
                f"/api/projects/{project_id}/students/{student_id}/render"
            )
            assert_status(rendered, 200)
        academic_term_id = reporting_term_id(
            client,
            [period_a["id"], period_b["id"]],
        )

        download = client.get(
            "/api/roster/semester-export/download",
            params={
                "academic_term_id": academic_term_id,
                "period_ids": [period_a["id"], period_b["id"]],
                "mode": "print",
            },
        )
        assert_status(download, 200)
        with ZipFile(BytesIO(download.content)) as zip_archive:
            pdf_paths = [
                name for name in zip_archive.namelist() if name.endswith(".pdf")
            ]
            merged_paths = [
                name for name in pdf_paths if "/全期合併_" in name
            ]
            assert len(pdf_paths) == 3
            assert merged_paths == [
                next(
                    name.rsplit("/", 1)[0]
                    for name in pdf_paths
                    if "/全期合併_" not in name
                ) + f"/全期合併_{child_name}.pdf"
            ]
            period_page_total = sum(
                len(PdfReader(BytesIO(zip_archive.read(name))).pages)
                for name in pdf_paths
                if name not in merged_paths
            )
            merged_reader = PdfReader(BytesIO(zip_archive.read(merged_paths[0])))
            assert period_page_total >= 2
            assert len(merged_reader.pages) == (period_page_total + 1) // 2
            for merged_page in merged_reader.pages:
                assert round(float(merged_page.mediabox.width)) == 1191
                assert round(float(merged_page.mediabox.height)) == 842
            manifest = zip_archive.read("匯出說明.txt").decode("utf-8")
        assert "全期合併" in manifest


def test_current_supervisor_reads_historical_snapshot_after_classroom_move():
    with started_client() as client:
        login(client)
        supervisor, supervisor_password = create_user(client, "supervisor")
        teacher, _ = create_user(client, "teacher")
        period = create_active_period(client)
        template_id, _ = create_template_with_page(client, period_id=period["id"])
        source_campus_id, classroom_id = create_scoped_classroom(client)
        target_campus_response = client.post(
            "/api/organization/campuses",
            json={"name": unique_name("moved_target")},
        )
        assert_status(target_campus_response, 201)
        target_campus_id = target_campus_response.json()["id"]
        set_campus_supervisor_scope(
            client,
            source_campus_id,
            supervisor["id"],
            department="infant",
        )
        set_classroom_teachers(client, classroom_id, [teacher["id"]])
        member_response = client.post(
            f"/api/organization/classrooms/{classroom_id}/members/batch",
            json={"members": [{"name": unique_name("history_child")}]},
        )
        assert_status(member_response, 201)
        member_id = member_response.json()["created"][0]["id"]
        project_id = create_classroom_project(client, classroom_id, template_id)
        academic_term_id = reporting_term_id(client, [period["id"]])

        ended_member = client.patch(
            f"/api/organization/classrooms/{classroom_id}/members/{member_id}",
            json={"status": "ended", "end_reason": "departed"},
        )
        assert_status(ended_member, 200)
        set_classroom_teachers(client, classroom_id, [])
        moved = client.patch(
            f"/api/organization/classrooms/{classroom_id}",
            json={"campus_id": target_campus_id, "is_active": False},
        )
        assert_status(moved, 200)

        client.cookies.clear()
        login(client, supervisor["username"], supervisor_password)
        detail = client.get(f"/api/projects/{project_id}")
        assert_status(detail, 200)
        progress = client.get(
            "/api/roster/teacher-progress",
            params={"academic_term_id": academic_term_id},
        )
        assert_status(progress, 200)
        assert _project_progress(progress.json(), project_id)["project_id"] == project_id

        client.cookies.clear()
        login(client)
        ended = client.put(
            f"/api/organization/campuses/{source_campus_id}/supervisors",
            json={
                "campus_supervisor_ids": [],
                "department_supervisors": [
                    {"department": "infant", "supervisor_ids": []},
                    {"department": "academy", "supervisor_ids": []},
                ],
            },
        )
        assert_status(ended, 200)

        client.cookies.clear()
        login(client, supervisor["username"], supervisor_password)
        forbidden_detail = client.get(f"/api/projects/{project_id}")
        assert_status(forbidden_detail, 403)
        forbidden_progress = client.get(
            "/api/roster/teacher-progress",
            params={"academic_term_id": academic_term_id},
        )
        assert_status(forbidden_progress, 403)


def test_duplicate_child_period_is_skipped_by_preview_render_and_download(monkeypatch):
    with started_client() as client:
        login(client)
        teacher, _ = create_user(client, "teacher")
        period = create_active_period(client)
        template_id, _ = create_template_with_page(client, period_id=period["id"])
        _, classroom_id = create_scoped_classroom(client)
        set_classroom_teachers(client, classroom_id, [teacher["id"]])
        student_name = unique_name("cross_project_duplicate")
        add_classroom_members(client, classroom_id, [student_name])
        project_id = create_classroom_project(client, classroom_id, template_id)
        student_id = add_students(client, project_id, [student_name])[student_name]
        academic_term_id = reporting_term_id(client, [period["id"]])

        db = SessionLocal()
        try:
            original_project = db.get(Project, project_id)
            original_student = db.get(Student, student_id)
            duplicate_project = Project(
                name=unique_name("legacy_duplicate"),
                template_id=original_project.template_id,
                department=original_project.department,
                template_period_id=original_project.template_period_id,
                template_revision=original_project.template_revision,
                owner_id=original_project.owner_id,
                classroom_id=original_project.classroom_id,
                class_period_work_slot_id=original_project.class_period_work_slot_id,
                created_by_id=original_project.created_by_id,
                created_by_name=original_project.created_by_name,
                campus_id_snapshot=original_project.campus_id_snapshot,
                campus_name_snapshot=original_project.campus_name_snapshot,
                classroom_name_snapshot=original_project.classroom_name_snapshot,
            )
            db.add(duplicate_project)
            db.flush()
            db.add(Student(
                project_id=duplicate_project.id,
                name=original_student.name,
                order_index=0,
                roster_child_id=original_student.roster_child_id,
            ))
            db.commit()
            duplicate_project_id = duplicate_project.id
        finally:
            db.close()

        preview = get_semester_preview(client, [period["id"]])
        assert_status(preview, 200)
        child = next(
            child
            for child in preview_children(preview.json())
            if child["name"] == student_name
        )
        cell = child["cells"][0]
        assert cell["status"] == "duplicate"
        assert {entry["project_id"] for entry in cell["entries"]} == {
            project_id,
            duplicate_project_id,
        }

        rendered_students = []
        monkeypatch.setattr(
            semester_render_service,
            "render_and_save_student_album",
            lambda *args: rendered_students.append(args[1].id),
        )
        db = SessionLocal()
        try:
            render_result = semester_render_service.render_missing_semester_albums(
                db,
                academic_term_id,
                [period["id"]],
            )
        finally:
            db.close()
        assert render_result["rendered"] == 0
        assert [error["code"] for error in render_result["errors"]] == ["duplicate"]
        assert rendered_students == []

        download = client.get(
            "/api/roster/semester-export/download",
            params={
                "academic_term_id": academic_term_id,
                "period_ids": [period["id"]],
                "mode": "print",
            },
        )
        assert_status(download, 200)
        with ZipFile(BytesIO(download.content)) as zip_archive:
            assert zip_archive.namelist() == ["匯出說明.txt"]
            manifest = zip_archive.read("匯出說明.txt").decode("utf-8")
        assert "同一期有重複相本" in manifest
        assert str(project_id) in manifest
        assert str(duplicate_project_id) in manifest
