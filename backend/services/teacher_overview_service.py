"""正式學期的班級 × 期別老師進度與 Excel 匯出。"""

import io
import json

from sqlalchemy.orm import Session, selectinload

from database import (
    AcademicTerm,
    AcademicTermClassroom,
    AcademicTermClassroomTeacher,
    AcademicTermPeriod,
    ClassPeriodWorkSlot,
    Project,
)
from services.label_texts import (
    get_label_entry_text,
    merge_project_label_texts_into_pages,
)
from services.layout_group_traversal import iter_layout_render_elements
from services.organization_scope_service import (
    OrganizationReadScope,
    REPORTING_TERM_STATUSES,
    apply_term_classroom_report_scope,
    load_reporting_term_or_404,
)
from services.student_render_service import get_template_page_layouts


def _text_label_is_fillable(label: dict) -> bool:
    """與前端 textLabelRoles.js 相同：固定文字不列入老師填寫進度。"""
    role = label.get("text_role", label.get("textRole"))
    return role != "static" and label.get("editable") is not False


def _summarize_student_progress(
    pages_data: list,
    page_layouts: list[dict],
    project_label_texts: dict,
) -> tuple[int, int, int, int]:
    """回傳單一學生的照片與老師可填文字之已填／總格數。"""
    merged_pages = merge_project_label_texts_into_pages(
        pages_data,
        project_label_texts,
        page_layouts,
    )
    merged_by_index = {
        page_data.get("page_index"): page_data
        for page_data in merged_pages
        if isinstance(page_data, dict)
    }

    photo_filled_count = 0
    photo_total_count = 0
    text_filled_count = 0
    text_total_count = 0
    for page_index, layout in enumerate(page_layouts):
        page_data = merged_by_index.get(page_index, {})
        if page_data.get("skip"):
            continue
        page_photos = page_data.get("photos") or {}
        merged_label_texts = page_data.get("label_texts") or {}
        for element_type, element, _ in iter_layout_render_elements(layout):
            if element_type == "photo":
                photo_total_count += 1
                if page_photos.get(str(element.get("id"))):
                    photo_filled_count += 1
                continue
            if element_type != "text" or not _text_label_is_fillable(element):
                continue
            text_total_count += 1
            label_id = str(element.get("id"))
            effective_text = get_label_entry_text(merged_label_texts.get(label_id))
            if str(effective_text or "").strip():
                text_filled_count += 1
    return (
        photo_filled_count,
        photo_total_count,
        text_filled_count,
        text_total_count,
    )


def _serialize_term_period(term_period: AcademicTermPeriod) -> dict:
    return {
        "id": term_period.template_period_id,
        "term_period_id": term_period.id,
        "template_period_id": term_period.template_period_id,
        "name": term_period.period_name_snapshot,
        "department": term_period.department,
        "position": term_period.position,
    }


def _serialize_term(term: AcademicTerm) -> dict:
    return {
        "id": term.id,
        "label": term.label,
        "status": term.status,
        "is_current": term.status in {"imported", "active"},
        "starts_on": term.starts_on.isoformat() if term.starts_on else None,
        "ends_on": term.ends_on.isoformat() if term.ends_on else None,
    }


def list_reporting_terms(
    db: Session,
    organization_scope: OrganizationReadScope,
) -> dict:
    """列出目前主管 scope 有學期班級的正式學期；admin 看全部。"""
    query = (
        db.query(AcademicTerm)
        .options(selectinload(AcademicTerm.periods))
        .filter(
            AcademicTerm.status.in_(REPORTING_TERM_STATUSES),
            AcademicTerm.periods.any(),
        )
    )
    visible_departments_by_term: dict[int, set[str]] | None = None
    if not organization_scope.is_admin:
        scoped_department_rows = apply_term_classroom_report_scope(
            db.query(
                AcademicTermClassroom.academic_term_id,
                AcademicTermClassroom.department,
            ),
            organization_scope,
        ).distinct().all()
        visible_departments_by_term = {}
        for academic_term_id, department in scoped_department_rows:
            visible_departments_by_term.setdefault(
                academic_term_id,
                set(),
            ).add(department)
        if not visible_departments_by_term:
            return {"terms": []}
        query = query.filter(
            AcademicTerm.id.in_(tuple(visible_departments_by_term))
        )
    terms = query.order_by(AcademicTerm.created_at.desc(), AcademicTerm.id.desc()).all()
    terms_payload = []
    for term in terms:
        visible_departments = (
            None
            if visible_departments_by_term is None
            else visible_departments_by_term.get(term.id, set())
        )
        periods = [
            _serialize_term_period(period)
            for period in sorted(term.periods, key=lambda row: row.position)
            if visible_departments is None
            or period.department in visible_departments
        ]
        if not periods:
            continue
        terms_payload.append({
            **_serialize_term(term),
            "periods": periods,
        })
    return {
        "terms": terms_payload
    }


def _load_report_classrooms(
    db: Session,
    academic_term_id: int,
    organization_scope: OrganizationReadScope,
    *,
    department: str | None = None,
    campus_id: int | None = None,
    classroom_id: int | None = None,
) -> list[AcademicTermClassroom]:
    slot_loader = selectinload(AcademicTermClassroom.work_slots)
    query = db.query(AcademicTermClassroom).options(
        selectinload(AcademicTermClassroom.teachers),
        slot_loader.selectinload(ClassPeriodWorkSlot.term_period),
        slot_loader.selectinload(ClassPeriodWorkSlot.projects).selectinload(
            Project.students
        ),
        slot_loader.selectinload(ClassPeriodWorkSlot.projects).selectinload(
            Project.owner
        ),
        slot_loader.selectinload(ClassPeriodWorkSlot.projects).selectinload(
            Project.template
        ),
    ).filter(AcademicTermClassroom.academic_term_id == academic_term_id)
    query = apply_term_classroom_report_scope(query, organization_scope)
    if department is not None:
        query = query.filter(AcademicTermClassroom.department == department)
    if campus_id is not None:
        query = query.filter(AcademicTermClassroom.campus_id_snapshot == campus_id)
    if classroom_id is not None:
        query = query.filter(AcademicTermClassroom.classroom_id == classroom_id)
    return query.order_by(
        AcademicTermClassroom.campus_name_snapshot,
        AcademicTermClassroom.classroom_name_snapshot,
        AcademicTermClassroom.id,
    ).all()


def _parse_json_list(value: str | None) -> list:
    try:
        parsed = json.loads(value or "[]")
    except ValueError:
        return []
    return parsed if isinstance(parsed, list) else []


def _parse_json_dict(value: str | None) -> dict:
    try:
        parsed = json.loads(value or "{}")
    except ValueError:
        return {}
    return parsed if isinstance(parsed, dict) else {}


def _serialize_project_progress(
    project: Project,
    page_layouts: list[dict],
) -> dict:
    project_label_texts = _parse_json_dict(project.label_texts_json)
    students_payload = []
    photo_filled = 0
    photo_total = 0
    text_filled = 0
    text_total = 0
    for student in project.students:
        (
            student_photo_filled,
            student_photo_total,
            student_text_filled,
            student_text_total,
        ) = _summarize_student_progress(
            _parse_json_list(student.pages_data_json),
            page_layouts,
            project_label_texts,
        )
        photo_filled += student_photo_filled
        photo_total += student_photo_total
        text_filled += student_text_filled
        text_total += student_text_total
        students_payload.append({
            "student_id": student.id,
            "student_name": student.name,
            "photo_filled": student_photo_filled,
            "photo_total": student_photo_total,
            "text_filled": student_text_filled,
            "text_total": student_text_total,
            "blank_text_count": student_text_total - student_text_filled,
        })

    student_count = len(students_payload)
    is_content_ready = (
        photo_filled == photo_total
        and text_filled == text_total
    )
    content_status = (
        "empty"
        if student_count == 0
        else "ready" if is_content_ready else "incomplete"
    )
    workflow_status = "submitted_locked" if project.completed_at else "working"

    attention_codes = []
    if content_status == "empty":
        attention_codes.append("empty_project")
    if project.completed_at and photo_filled < photo_total:
        attention_codes.append("submitted_with_missing_photos")
    if project.completed_at and text_filled < text_total:
        attention_codes.append("submitted_with_missing_texts")

    return {
        "project_id": project.id,
        "project_name": project.name,
        "owner_id": project.owner_id,
        "owner_name": project.owner.display_name if project.owner else None,
        "student_count": student_count,
        "photo_filled": photo_filled,
        "photo_total": photo_total,
        "text_filled": text_filled,
        "text_total": text_total,
        "blank_text_count": text_total - text_filled,
        "content_status": content_status,
        "workflow_status": workflow_status,
        "attention_codes": attention_codes,
        "completed_at": (
            project.completed_at.isoformat() if project.completed_at else None
        ),
        "students": students_payload,
    }


def _serialize_teacher(teacher: AcademicTermClassroomTeacher) -> dict:
    return {
        "user_id": teacher.teacher_id,
        "display_name": teacher.teacher_name_snapshot,
        "duty": teacher.duty,
    }


def build_teacher_progress_overview(
    db: Session,
    academic_term_id: int,
    organization_scope: OrganizationReadScope,
    *,
    department: str | None = None,
    campus_id: int | None = None,
    classroom_id: int | None = None,
) -> dict:
    """以正式工作格建立班級 × 期別進度，不按 owner 複製工作。"""
    term = load_reporting_term_or_404(
        db,
        academic_term_id,
        organization_scope,
    )
    term_classrooms = _load_report_classrooms(
        db,
        academic_term_id,
        organization_scope,
        department=department,
        campus_id=campus_id,
        classroom_id=classroom_id,
    )
    report_periods = [
        period
        for period in sorted(term.periods, key=lambda row: row.position)
        if department is None or period.department == department
    ]
    allowed_term_period_ids = {period.id for period in report_periods}
    layouts_by_template: dict[int, list[dict]] = {}

    classrooms_payload = []
    summary = {
        "classroom_count": len(term_classrooms),
        "slot_count": 0,
        "not_created_slot_count": 0,
        "archived_slot_count": 0,
        "single_slot_count": 0,
        "multiple_projects_slot_count": 0,
        "project_count": 0,
        "content_ready_project_count": 0,
        "submitted_project_count": 0,
        "attention_project_count": 0,
    }
    for term_classroom in term_classrooms:
        slots_payload = []
        for slot in sorted(
            term_classroom.work_slots,
            key=lambda row: (row.term_period.position, row.id),
        ):
            if slot.term_period_id not in allowed_term_period_ids:
                continue
            active_projects = [
                project for project in slot.projects if project.deleted_at is None
            ]
            if not active_projects:
                creation_status = "archived" if slot.started_at else "not_created"
            elif len(active_projects) == 1:
                creation_status = "single"
            else:
                creation_status = "multiple_projects"
            projects_payload = []
            for project in active_projects:
                if project.template_id not in layouts_by_template:
                    layouts_by_template[project.template_id] = (
                        get_template_page_layouts(project)
                    )
                project_payload = _serialize_project_progress(
                    project,
                    layouts_by_template[project.template_id],
                )
                projects_payload.append(project_payload)
                summary["project_count"] += 1
                summary["content_ready_project_count"] += int(
                    project_payload["content_status"] == "ready"
                )
                summary["submitted_project_count"] += int(
                    project_payload["workflow_status"] == "submitted_locked"
                )
                summary["attention_project_count"] += int(
                    bool(project_payload["attention_codes"])
                )
            summary["slot_count"] += 1
            summary[f"{creation_status}_slot_count"] += 1
            slots_payload.append({
                "work_slot_id": slot.id,
                "term_period_id": slot.term_period_id,
                "period_id": slot.term_period.template_period_id,
                "template_period_id": slot.term_period.template_period_id,
                "period_name": slot.term_period.period_name_snapshot,
                "position": slot.term_period.position,
                "started_at": (
                    slot.started_at.isoformat() if slot.started_at else None
                ),
                "creation_status": creation_status,
                "projects": projects_payload,
            })
        classrooms_payload.append({
            "term_classroom_id": term_classroom.id,
            "classroom_id": term_classroom.classroom_id,
            "campus_id": term_classroom.campus_id_snapshot,
            "campus_name": term_classroom.campus_name_snapshot,
            "classroom_name": term_classroom.classroom_name_snapshot,
            "department": term_classroom.department,
            "teachers": [
                _serialize_teacher(teacher)
                for teacher in sorted(
                    term_classroom.teachers,
                    key=lambda row: (row.duty != "lead", row.id),
                )
            ],
            "slots": slots_payload,
        })

    return {
        "term": _serialize_term(term),
        "periods": [_serialize_term_period(period) for period in report_periods],
        "summary": summary,
        "classrooms": classrooms_payload,
    }


def build_teacher_overview_workbook(
    db: Session,
    academic_term_id: int,
    organization_scope: OrganizationReadScope,
    *,
    department: str | None = None,
    campus_id: int | None = None,
    classroom_id: int | None = None,
) -> bytes:
    """輸出與畫面同一資料來源的摘要、班級期別與學生明細。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    overview = build_teacher_progress_overview(
        db,
        academic_term_id,
        organization_scope,
        department=department,
        campus_id=campus_id,
        classroom_id=classroom_id,
    )
    workbook = Workbook()
    header_font = Font(bold=True)

    def excel_safe(value):
        """避免使用者文字在 Excel 被解析為公式。"""
        if not isinstance(value, str):
            return value
        if value.lstrip().startswith(("=", "+", "-", "@")):
            return f"'{value}"
        return value

    def append_safe(sheet, values) -> None:
        sheet.append([excel_safe(value) for value in values])

    summary_sheet = workbook.active
    summary_sheet.title = "摘要"
    append_safe(summary_sheet, ["項目", "數量"])
    summary_labels = {
        "classroom_count": "班級數",
        "slot_count": "工作格數",
        "not_created_slot_count": "未建立工作格",
        "archived_slot_count": "已封存工作格",
        "single_slot_count": "單一專案工作格",
        "multiple_projects_slot_count": "多專案工作格",
        "project_count": "專案數",
        "content_ready_project_count": "照片與文字內容已齊專案",
        "submitted_project_count": "已交件鎖定專案",
        "attention_project_count": "需注意專案",
    }
    for summary_key, summary_label in summary_labels.items():
        append_safe(
            summary_sheet,
            [summary_label, overview["summary"][summary_key]],
        )

    slot_sheet = workbook.create_sheet("班級期別")
    append_safe(slot_sheet, [
        "分校",
        "部門",
        "班級",
        "期別",
        "主教",
        "協同老師",
        "建立狀態",
        "專案數",
        "負責人",
        "工作流",
        "相本學生數",
        "照片已填",
        "照片總格",
        "文字已填",
        "文字總格",
        "內容狀態",
        "空白文字格",
        "異常",
    ])
    student_sheet = workbook.create_sheet("學生明細")
    append_safe(student_sheet, [
        "分校",
        "部門",
        "班級",
        "期別",
        "專案 ID",
        "專案",
        "負責人",
        "學生",
        "照片已填",
        "照片總格",
        "文字已填",
        "文字總格",
        "空白文字格",
        "工作流",
    ])

    for classroom in overview["classrooms"]:
        lead_names = [
            teacher["display_name"]
            for teacher in classroom["teachers"]
            if teacher["duty"] == "lead"
        ]
        co_teacher_names = [
            teacher["display_name"]
            for teacher in classroom["teachers"]
            if teacher["duty"] == "co_teacher"
        ]
        for slot in classroom["slots"]:
            projects = slot["projects"]
            append_safe(slot_sheet, [
                classroom["campus_name"],
                classroom["department"],
                classroom["classroom_name"],
                slot["period_name"],
                "、".join(lead_names),
                "、".join(co_teacher_names),
                slot["creation_status"],
                len(projects),
                "、".join(
                    project["owner_name"] or "未指定" for project in projects
                ),
                "、".join(project["workflow_status"] for project in projects),
                sum(project["student_count"] for project in projects),
                sum(project["photo_filled"] for project in projects),
                sum(project["photo_total"] for project in projects),
                sum(project["text_filled"] for project in projects),
                sum(project["text_total"] for project in projects),
                "、".join(project["content_status"] for project in projects),
                sum(project["blank_text_count"] for project in projects),
                "、".join(
                    code
                    for project in projects
                    for code in project["attention_codes"]
                ),
            ])
            for project in projects:
                for student in project["students"]:
                    append_safe(student_sheet, [
                        classroom["campus_name"],
                        classroom["department"],
                        classroom["classroom_name"],
                        slot["period_name"],
                        project["project_id"],
                        project["project_name"],
                        project["owner_name"],
                        student["student_name"],
                        student["photo_filled"],
                        student["photo_total"],
                        student["text_filled"],
                        student["text_total"],
                        student["blank_text_count"],
                        project["workflow_status"],
                    ])

    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = header_font
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(
                max_length * 2 + 4,
                40,
            )

    output_buffer = io.BytesIO()
    workbook.save(output_buffer)
    return output_buffer.getvalue()
